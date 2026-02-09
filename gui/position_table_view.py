"""
CANSLIM Monitor - Position Table View
Spreadsheet-style view of all positions with Excel-like column filtering and editing.
"""

from datetime import date, datetime
from typing import Optional, Dict, Any, List, Set

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableView, QLineEdit,
    QPushButton, QLabel, QHeaderView, QComboBox, QWidget,
    QAbstractItemView, QMessageBox, QFrame, QGroupBox, QFormLayout,
    QScrollArea, QCheckBox, QListWidget, QListWidgetItem, QMenu,
    QWidgetAction, QSpinBox, QDoubleSpinBox, QInputDialog, QFileDialog
)
from PyQt6.QtCore import (
    Qt, QAbstractTableModel, QModelIndex, QSortFilterProxyModel,
    pyqtSignal, QTimer, QPoint
)
from PyQt6.QtGui import QFont, QColor, QBrush, QAction

import yaml
import os
import csv

from canslim_monitor.gui.state_config import STATES


# Define column types for filtering behavior
COLUMN_TYPES = {
    'symbol': 'text',
    'portfolio': 'text',
    'state': 'text',
    'pattern': 'text',
    'pivot': 'float',
    'last_price': 'float',
    'dist_from_pivot': 'float',  # Computed: ((last_price - pivot) / pivot) * 100
    'unrealized_pnl': 'float',   # Computed: (last_price - avg_cost) * total_shares
    'unrealized_pct': 'float',   # Computed: ((last_price - avg_cost) / avg_cost) * 100
    'realized_pnl': 'float',     # From DB: Realized P&L $
    'total_pnl': 'float',        # Computed: realized + unrealized
    'rs_rating': 'int',
    'rs_3mo': 'int',
    'rs_6mo': 'int',
    'eps_rating': 'int',
    'comp_rating': 'int',
    'smr_rating': 'text',
    'ad_rating': 'text',
    'industry_rank': 'int',
    'base_stage': 'text',
    'base_depth': 'float',
    'base_length': 'int',
    'ud_vol_ratio': 'float',
    'prior_uptrend': 'float',
    'fund_count': 'int',
    'funds_qtr_chg': 'int',
    'entry_grade': 'text',
    'entry_score': 'int',
    'watch_date': 'date',
    'entry_date': 'date',
    'breakout_date': 'date',
    'earnings_date': 'date',
    'total_shares': 'int',
    'avg_cost': 'float',
    'hard_stop_pct': 'float',
    'stop_price': 'float',
    'health_score': 'int',
    # Closed position fields
    'close_price': 'float',
    'close_date': 'date',
    'close_reason': 'text',
    'realized_pnl_pct': 'float',
    # State -1.5 (WATCHING_EXITED) fields
    'original_pivot': 'float',
    'ma_test_count': 'int',
    'watching_exited_since': 'date',
    'notes': 'text',
}


class ColumnVisibilityDialog(QDialog):
    """Dialog for selecting which columns to show/hide."""

    def __init__(self, columns: list, hidden_columns: set, parent=None):
        super().__init__(parent)
        self.columns = columns  # List of (key, name) tuples
        self.hidden_columns = set(hidden_columns)  # Copy to work with
        self._checkboxes: Dict[int, QCheckBox] = {}

        self.setWindowTitle("Select Columns")
        self.setMinimumWidth(280)
        self.setMaximumHeight(600)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        self.setStyleSheet("""
            QDialog {
                background-color: #2D2D30;
            }
            QLabel {
                color: #E0E0E0;
            }
            QCheckBox {
                color: #E0E0E0;
                padding: 4px 8px;
                font-size: 12px;
            }
            QCheckBox:hover {
                background-color: #3C3C3F;
                border-radius: 3px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
            QCheckBox::indicator:checked {
                background-color: #0078D4;
                border: 1px solid #0078D4;
                border-radius: 3px;
            }
            QCheckBox::indicator:unchecked {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                border-radius: 3px;
            }
            QPushButton {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                border-radius: 4px;
                padding: 6px 12px;
                color: #E0E0E0;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #4A4A4D;
                border-color: #0078D4;
            }
            QScrollArea {
                border: 1px solid #4A4A4D;
                border-radius: 4px;
                background-color: #2D2D30;
            }
        """)

        # Title
        title = QLabel("Select Columns to Display")
        title.setStyleSheet("font-weight: bold; font-size: 13px; padding: 2px;")
        layout.addWidget(title)

        # Select All / Select None buttons
        btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("✓ Select All")
        select_all_btn.clicked.connect(self._select_all)
        btn_layout.addWidget(select_all_btn)

        select_none_btn = QPushButton("✗ Select None")
        select_none_btn.clicked.connect(self._select_none)
        btn_layout.addWidget(select_none_btn)
        layout.addLayout(btn_layout)

        # Scrollable checkbox list
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(400)

        list_widget = QWidget()
        list_widget.setStyleSheet("background-color: #2D2D30;")
        list_layout = QVBoxLayout(list_widget)
        list_layout.setContentsMargins(4, 4, 4, 4)
        list_layout.setSpacing(2)

        for idx, (key, name) in enumerate(self.columns):
            cb = QCheckBox(name)
            cb.setChecked(idx not in self.hidden_columns)
            self._checkboxes[idx] = cb
            list_layout.addWidget(cb)

        list_layout.addStretch()
        scroll.setWidget(list_widget)
        layout.addWidget(scroll)

        # Apply / Cancel buttons
        action_layout = QHBoxLayout()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        action_layout.addWidget(cancel_btn)

        action_layout.addStretch()

        apply_btn = QPushButton("Apply")
        apply_btn.clicked.connect(self.accept)
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        action_layout.addWidget(apply_btn)
        layout.addLayout(action_layout)

    def _select_all(self):
        """Check all checkboxes."""
        for cb in self._checkboxes.values():
            cb.setChecked(True)

    def _select_none(self):
        """Uncheck all checkboxes except Symbol (index 0)."""
        for idx, cb in self._checkboxes.items():
            if idx == 0:  # Keep Symbol visible
                cb.setChecked(True)
            else:
                cb.setChecked(False)

    def get_hidden_columns(self) -> set:
        """Return set of column indices that should be hidden."""
        hidden = set()
        for idx, cb in self._checkboxes.items():
            if not cb.isChecked():
                hidden.add(idx)
        return hidden


class TextFilterPopup(QDialog):
    """Popup for filtering text columns with multi-select checkboxes."""

    filter_changed = pyqtSignal(set)  # Emits set of selected values
    
    def __init__(self, unique_values: List[str], selected_values: Set[str] = None, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setMinimumWidth(220)
        self.setMaximumHeight(400)
        
        self.unique_values = sorted([v for v in unique_values if v], key=lambda x: str(x).lower())
        self.selected_values = selected_values or set()
        
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #FFFFFF;
                border: 2px solid #0078D4;
                border-radius: 6px;
            }
            QCheckBox {
                color: #333333;
                padding: 4px 8px;
                font-size: 12px;
            }
            QCheckBox:hover {
                background-color: #E8F4FD;
                border-radius: 3px;
            }
            QPushButton {
                background-color: #F0F0F0;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 5px 10px;
                color: #333333;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
                border-color: #0078D4;
            }
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px;
                color: #333333;
                font-size: 12px;
            }
            QLineEdit:focus {
                border: 2px solid #0078D4;
            }
            QScrollArea {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
            }
        """)
        
        # Title
        title = QLabel("Select Values")
        title.setStyleSheet("font-weight: bold; font-size: 13px; color: #333333; padding: 2px;")
        layout.addWidget(title)
        
        # Search box
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Search...")
        self.search_input.textChanged.connect(self._filter_list)
        layout.addWidget(self.search_input)
        
        # Select All / Clear buttons
        btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("✓ Select All")
        select_all_btn.clicked.connect(self._select_all)
        btn_layout.addWidget(select_all_btn)
        
        clear_btn = QPushButton("✗ Clear All")
        clear_btn.clicked.connect(self._clear_all)
        btn_layout.addWidget(clear_btn)
        layout.addLayout(btn_layout)
        
        # Scrollable checkbox list
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(220)
        
        self.list_widget = QWidget()
        self.list_widget.setStyleSheet("background-color: #FFFFFF;")
        self.list_layout = QVBoxLayout(self.list_widget)
        self.list_layout.setContentsMargins(4, 4, 4, 4)
        self.list_layout.setSpacing(2)
        
        self.checkboxes: Dict[str, QCheckBox] = {}
        all_selected = len(self.selected_values) == 0
        
        for value in self.unique_values:
            cb = QCheckBox(str(value))
            cb.setChecked(all_selected or value in self.selected_values)
            self.checkboxes[value] = cb
            self.list_layout.addWidget(cb)
        
        self.list_layout.addStretch()
        scroll.setWidget(self.list_widget)
        layout.addWidget(scroll)
        
        # Apply button
        apply_btn = QPushButton("Apply Filter")
        apply_btn.clicked.connect(self._apply_filter)
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        layout.addWidget(apply_btn)
    
    def _filter_list(self, text: str):
        """Filter the checkbox list based on search text."""
        text = text.lower()
        for value, cb in self.checkboxes.items():
            cb.setVisible(text in str(value).lower())
    
    def _select_all(self):
        """Select all visible checkboxes."""
        for cb in self.checkboxes.values():
            if cb.isVisible():
                cb.setChecked(True)
    
    def _clear_all(self):
        """Clear all visible checkboxes."""
        for cb in self.checkboxes.values():
            if cb.isVisible():
                cb.setChecked(False)
    
    def _apply_filter(self):
        """Apply the filter and close."""
        selected = {value for value, cb in self.checkboxes.items() if cb.isChecked()}
        self.filter_changed.emit(selected)
        self.close()


class NumericFilterPopup(QDialog):
    """Popup for filtering numeric columns with comparison operators."""
    
    filter_changed = pyqtSignal(dict)  # Emits {'operator': str, 'value': float, 'value2': float}
    
    def __init__(self, column_type: str = 'float', current_filter: Dict = None, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setFixedWidth(280)
        
        self.column_type = column_type
        self.current_filter = current_filter or {}
        
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #FFFFFF;
                border: 2px solid #0078D4;
                border-radius: 6px;
            }
            QLabel {
                color: #333333;
                font-size: 12px;
            }
            QComboBox {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px;
                color: #333333;
                font-size: 12px;
                min-height: 24px;
            }
            QComboBox:focus {
                border: 2px solid #0078D4;
            }
            QComboBox::drop-down {
                border: none;
                width: 24px;
            }
            QComboBox QAbstractItemView {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                selection-background-color: #0078D4;
                selection-color: white;
            }
            QSpinBox, QDoubleSpinBox {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px;
                color: #333333;
                font-size: 12px;
                min-height: 24px;
            }
            QSpinBox:focus, QDoubleSpinBox:focus {
                border: 2px solid #0078D4;
            }
            QPushButton {
                background-color: #F0F0F0;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px 12px;
                color: #333333;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
                border-color: #0078D4;
            }
        """)
        
        # Title
        title = QLabel("Numeric Filter")
        title.setStyleSheet("font-weight: bold; font-size: 13px; padding: 2px;")
        layout.addWidget(title)
        
        # Operator selection
        op_layout = QFormLayout()
        self.operator_combo = QComboBox()
        self.operator_combo.addItems([
            "(No Filter)",
            "=  Equals",
            "≠  Not Equals", 
            ">  Greater Than",
            "≥  Greater or Equal",
            "<  Less Than",
            "≤  Less or Equal",
            "↔  Between"
        ])
        self.operator_combo.currentIndexChanged.connect(self._on_operator_changed)
        op_layout.addRow("Condition:", self.operator_combo)
        layout.addLayout(op_layout)
        
        # Value input(s)
        value_layout = QFormLayout()
        
        if self.column_type == 'int':
            self.value_input = QSpinBox()
            self.value_input.setRange(-999999, 999999)
            self.value2_input = QSpinBox()
            self.value2_input.setRange(-999999, 999999)
        else:
            self.value_input = QDoubleSpinBox()
            self.value_input.setRange(-999999, 999999)
            self.value_input.setDecimals(2)
            self.value2_input = QDoubleSpinBox()
            self.value2_input.setRange(-999999, 999999)
            self.value2_input.setDecimals(2)
        
        value_layout.addRow("Value:", self.value_input)
        
        # Second value for "Between"
        self.value2_label = QLabel("And:")
        value_layout.addRow(self.value2_label, self.value2_input)
        self.value2_label.setVisible(False)
        self.value2_input.setVisible(False)
        
        layout.addLayout(value_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        clear_btn = QPushButton("Clear Filter")
        clear_btn.clicked.connect(self._clear_filter)
        btn_layout.addWidget(clear_btn)
        
        apply_btn = QPushButton("Apply")
        apply_btn.clicked.connect(self._apply_filter)
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        btn_layout.addWidget(apply_btn)
        layout.addLayout(btn_layout)
        
        # Load current filter
        if self.current_filter:
            op = self.current_filter.get('operator', '')
            operators = ["", "=", "≠", ">", "≥", "<", "≤", "between"]
            if op in operators:
                self.operator_combo.setCurrentIndex(operators.index(op))
            self.value_input.setValue(self.current_filter.get('value', 0))
            self.value2_input.setValue(self.current_filter.get('value2', 0))
            self._on_operator_changed(self.operator_combo.currentIndex())
    
    def _on_operator_changed(self, index: int):
        """Show/hide second value input for 'Between' operator."""
        is_between = index == 7  # "Between" option
        is_no_filter = index == 0
        
        self.value2_label.setVisible(is_between)
        self.value2_input.setVisible(is_between)
        self.value_input.setEnabled(not is_no_filter)
    
    def _clear_filter(self):
        """Clear the filter."""
        self.filter_changed.emit({})
        self.close()
    
    def _apply_filter(self):
        """Apply the filter."""
        operators = ["", "=", "≠", ">", "≥", "<", "≤", "between"]
        op = operators[self.operator_combo.currentIndex()]
        
        if not op:
            self.filter_changed.emit({})
        else:
            result = {
                'operator': op,
                'value': self.value_input.value(),
            }
            if op == 'between':
                result['value2'] = self.value2_input.value()
            self.filter_changed.emit(result)
        
        self.close()


class DateFilterPopup(QDialog):
    """Popup for filtering date columns with comparison operators."""

    filter_changed = pyqtSignal(dict)  # Emits {'operator': str, 'value': date, 'value2': date}

    def __init__(self, current_filter: Dict = None, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setFixedWidth(300)

        self.current_filter = current_filter or {}
        self._setup_ui()

    def _setup_ui(self):
        from PyQt6.QtWidgets import QDateEdit
        from PyQt6.QtCore import QDate

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        self.setStyleSheet("""
            QDialog {
                background-color: #FFFFFF;
                border: 2px solid #0078D4;
                border-radius: 6px;
            }
            QLabel {
                color: #333333;
                font-size: 12px;
            }
            QComboBox {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px;
                color: #333333;
                font-size: 12px;
                min-height: 24px;
            }
            QComboBox:focus {
                border: 2px solid #0078D4;
            }
            QComboBox::drop-down {
                border: none;
                width: 24px;
            }
            QComboBox QAbstractItemView {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                selection-background-color: #0078D4;
                selection-color: white;
            }
            QDateEdit {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px;
                color: #333333;
                font-size: 12px;
                min-height: 24px;
            }
            QDateEdit:focus {
                border: 2px solid #0078D4;
            }
            QPushButton {
                background-color: #F0F0F0;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px 12px;
                color: #333333;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
                border-color: #0078D4;
            }
        """)

        # Title
        title = QLabel("Date Filter")
        title.setStyleSheet("font-weight: bold; font-size: 13px; padding: 2px;")
        layout.addWidget(title)

        # Operator selection
        op_layout = QFormLayout()
        self.operator_combo = QComboBox()
        self.operator_combo.addItems([
            "(No Filter)",
            "=  On Date",
            ">  After",
            "≥  On or After",
            "<  Before",
            "≤  On or Before",
            "↔  Between"
        ])
        self.operator_combo.currentIndexChanged.connect(self._on_operator_changed)
        op_layout.addRow("Condition:", self.operator_combo)
        layout.addLayout(op_layout)

        # Date input(s)
        value_layout = QFormLayout()

        self.value_input = QDateEdit()
        self.value_input.setCalendarPopup(True)
        self.value_input.setDate(QDate.currentDate())
        self.value_input.setDisplayFormat("yyyy-MM-dd")
        value_layout.addRow("Date:", self.value_input)

        # Second date for "Between"
        self.value2_label = QLabel("And:")
        self.value2_input = QDateEdit()
        self.value2_input.setCalendarPopup(True)
        self.value2_input.setDate(QDate.currentDate())
        self.value2_input.setDisplayFormat("yyyy-MM-dd")
        value_layout.addRow(self.value2_label, self.value2_input)
        self.value2_label.setVisible(False)
        self.value2_input.setVisible(False)

        layout.addLayout(value_layout)

        # Buttons
        btn_layout = QHBoxLayout()

        clear_btn = QPushButton("Clear Filter")
        clear_btn.clicked.connect(self._clear_filter)
        btn_layout.addWidget(clear_btn)

        apply_btn = QPushButton("Apply")
        apply_btn.clicked.connect(self._apply_filter)
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        btn_layout.addWidget(apply_btn)
        layout.addLayout(btn_layout)

        # Load current filter
        if self.current_filter:
            op = self.current_filter.get('operator', '')
            operators = ["", "=", ">", "≥", "<", "≤", "between"]
            if op in operators:
                self.operator_combo.setCurrentIndex(operators.index(op))

            val = self.current_filter.get('value')
            if val:
                if isinstance(val, date):
                    self.value_input.setDate(QDate(val.year, val.month, val.day))

            val2 = self.current_filter.get('value2')
            if val2:
                if isinstance(val2, date):
                    self.value2_input.setDate(QDate(val2.year, val2.month, val2.day))

            self._on_operator_changed(self.operator_combo.currentIndex())

    def _on_operator_changed(self, index: int):
        """Show/hide second date input for 'Between' operator."""
        is_between = index == 6  # "Between" option
        is_no_filter = index == 0

        self.value2_label.setVisible(is_between)
        self.value2_input.setVisible(is_between)
        self.value_input.setEnabled(not is_no_filter)

    def _clear_filter(self):
        """Clear the filter."""
        self.filter_changed.emit({})
        self.close()

    def _apply_filter(self):
        """Apply the filter."""
        operators = ["", "=", ">", "≥", "<", "≤", "between"]
        op = operators[self.operator_combo.currentIndex()]

        if not op:
            self.filter_changed.emit({})
        else:
            qdate = self.value_input.date()
            result = {
                'operator': op,
                'value': date(qdate.year(), qdate.month(), qdate.day()),
                'type': 'date'
            }
            if op == 'between':
                qdate2 = self.value2_input.date()
                result['value2'] = date(qdate2.year(), qdate2.month(), qdate2.day())
            self.filter_changed.emit(result)

        self.close()


class ColumnFilterButton(QPushButton):
    """Button that shows filter status and opens filter popup."""
    
    filter_changed = pyqtSignal(int, object)  # column_index, filter_value
    
    def __init__(self, column_index: int, column_key: str, column_name: str, parent=None):
        super().__init__(parent)
        self.column_index = column_index
        self.column_key = column_key
        self.column_name = column_name
        self.column_type = COLUMN_TYPES.get(column_key, 'text')
        
        self.current_filter = None
        self.unique_values: List[str] = []
        
        self._update_appearance()
        self.clicked.connect(self._show_popup)
    
    def _update_appearance(self):
        """Update button appearance based on filter state."""
        if self.current_filter:
            self.setText("▼")
            self.setStyleSheet("""
                QPushButton {
                    background-color: #0078D4;
                    color: white;
                    border: none;
                    border-radius: 2px;
                    font-size: 9px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #106EBE;
                }
            """)
            
            # Build tooltip showing active filter
            if isinstance(self.current_filter, set):
                count = len(self.current_filter)
                tip = f"{self.column_name}: {count} value(s) selected"
            elif isinstance(self.current_filter, dict):
                op = self.current_filter.get('operator', '')
                val = self.current_filter.get('value', 0)
                if op == 'between':
                    val2 = self.current_filter.get('value2', 0)
                    tip = f"{self.column_name}: {val} to {val2}"
                else:
                    tip = f"{self.column_name}: {op} {val}"
            else:
                tip = f"{self.column_name}: Filter active"
            self.setToolTip(tip)
        else:
            self.setText("▼")
            self.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    color: #888888;
                    border: none;
                    font-size: 9px;
                }
                QPushButton:hover {
                    background-color: #E0E0E0;
                    color: #333333;
                    border-radius: 2px;
                }
            """)
            self.setToolTip(f"Filter {self.column_name}")
    
    def set_unique_values(self, values: List[str]):
        """Set the unique values for text filter popup."""
        self.unique_values = values
    
    def _show_popup(self):
        """Show the appropriate filter popup."""
        if self.column_type in ('int', 'float'):
            popup = NumericFilterPopup(
                self.column_type,
                self.current_filter if isinstance(self.current_filter, dict) else None,
                self
            )
            popup.filter_changed.connect(self._on_numeric_filter)
        elif self.column_type == 'date':
            popup = DateFilterPopup(
                self.current_filter if isinstance(self.current_filter, dict) else None,
                self
            )
            popup.filter_changed.connect(self._on_date_filter)
        else:
            # Text filter
            selected = self.current_filter if isinstance(self.current_filter, set) else set()
            popup = TextFilterPopup(
                self.unique_values,
                selected,
                self
            )
            popup.filter_changed.connect(self._on_text_filter)

        # Position popup below the button
        pos = self.mapToGlobal(QPoint(0, self.height()))
        popup.move(pos)
        popup.show()
    
    def _on_text_filter(self, selected_values: Set[str]):
        """Handle text filter selection."""
        if len(selected_values) == len(self.unique_values) or len(selected_values) == 0:
            self.current_filter = None
        else:
            self.current_filter = selected_values
        
        self._update_appearance()
        self.filter_changed.emit(self.column_index, self.current_filter)
    
    def _on_numeric_filter(self, filter_dict: Dict):
        """Handle numeric filter selection."""
        if not filter_dict:
            self.current_filter = None
        else:
            self.current_filter = filter_dict

        self._update_appearance()
        self.filter_changed.emit(self.column_index, self.current_filter)

    def _on_date_filter(self, filter_dict: Dict):
        """Handle date filter selection."""
        if not filter_dict:
            self.current_filter = None
        else:
            self.current_filter = filter_dict

        self._update_appearance()
        self.filter_changed.emit(self.column_index, self.current_filter)

    def clear_filter(self):
        """Clear the filter."""
        self.current_filter = None
        self._update_appearance()


class PositionTableModel(QAbstractTableModel):
    """Table model for displaying positions in a spreadsheet view."""
    
    COLUMNS = [
        ('symbol', 'Symbol'),
        ('portfolio', 'Portfolio'),
        ('state', 'State'),
        ('pattern', 'Pattern'),
        ('pivot', 'Pivot'),
        ('last_price', 'Last Price'),
        ('dist_from_pivot', 'Dist %'),      # Computed: distance from pivot
        ('unrealized_pct', 'Unreal %'),     # Computed: unrealized P&L %
        ('unrealized_pnl', 'Unreal $'),     # Computed: unrealized P&L $
        ('realized_pnl', 'Real $'),         # From DB: realized P&L $
        ('total_pnl', 'Total $'),           # Computed: realized + unrealized
        ('rs_rating', 'RS'),
        ('rs_3mo', 'RS 3M'),
        ('rs_6mo', 'RS 6M'),
        ('eps_rating', 'EPS'),
        ('comp_rating', 'Comp'),
        ('smr_rating', 'SMR'),
        ('ad_rating', 'A/D'),
        ('industry_rank', 'Ind Rank'),
        ('base_stage', 'Stage'),
        ('base_depth', 'Depth %'),
        ('base_length', 'Length'),
        ('ud_vol_ratio', 'U/D Vol'),        # Up/Down volume ratio
        ('prior_uptrend', 'Prior Up'),      # Prior uptrend %
        ('fund_count', 'Funds'),
        ('funds_qtr_chg', 'Funds Chg'),
        ('entry_grade', 'Grade'),
        ('entry_score', 'Score'),
        ('watch_date', 'Watch Date'),
        ('entry_date', 'Entry Date'),       # When position was entered
        ('breakout_date', 'Breakout'),
        ('earnings_date', 'Earnings'),      # Next earnings date
        ('total_shares', 'Shares'),
        ('avg_cost', 'Avg Cost'),
        ('hard_stop_pct', 'Stop %'),        # Hard stop percentage
        ('stop_price', 'Stop $'),           # Calculated stop price
        ('health_score', 'Health'),         # Position health score
        # Closed position fields
        ('close_price', 'Close $'),         # Price when closed
        ('close_date', 'Close Date'),       # Date position was closed
        ('close_reason', 'Close Reason'),   # Reason for closing
        ('realized_pnl_pct', 'Real %'),     # Realized P&L percentage
        # State -1.5 (WATCHING_EXITED) fields
        ('original_pivot', 'Orig Pivot'),   # Preserved pivot for retest detection
        ('ma_test_count', 'MA Tests'),      # Number of MA bounce tests (max 3)
        ('watching_exited_since', 'Watch Exit'),  # When entered State -1.5
        ('notes', 'Notes'),                 # Position notes
    ]
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._data: List[Dict[str, Any]] = []
        self._column_keys = [col[0] for col in self.COLUMNS]
        self._column_names = [col[1] for col in self.COLUMNS]
    
    def load_positions(self, positions: List[Dict[str, Any]]):
        """Load position data into the model."""
        self.beginResetModel()
        self._data = positions
        self.endResetModel()
    
    def rowCount(self, parent=QModelIndex()) -> int:
        return len(self._data)
    
    def columnCount(self, parent=QModelIndex()) -> int:
        return len(self.COLUMNS)
    
    def get_unique_values(self, column: int) -> List[str]:
        """Get unique values for a column (for text filter popup)."""
        col_key = self._column_keys[column]
        values = set()
        for row_data in self._data:
            value = row_data.get(col_key)
            if col_key == 'state':
                state_info = STATES.get(value)
                display = state_info.display_name if state_info else str(value)
                values.add(display)
            elif value is not None:
                values.add(str(value))
        return list(values)
    
    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or index.row() >= len(self._data):
            return None
        
        row_data = self._data[index.row()]
        col_key = self._column_keys[index.column()]
        value = row_data.get(col_key)
        
        if role == Qt.ItemDataRole.DisplayRole:
            if col_key == 'state':
                state_info = STATES.get(value)
                return state_info.display_name if state_info else str(value)
            elif col_key in ('pivot', 'last_price', 'avg_cost', 'stop_price', 'close_price', 'original_pivot'):
                return f"${value:.2f}" if value else ""
            elif col_key in ('dist_from_pivot', 'unrealized_pct', 'realized_pnl_pct', 'base_depth', 'hard_stop_pct', 'prior_uptrend'):
                return f"{value:+.1f}%" if value is not None else ""
            elif col_key in ('unrealized_pnl', 'realized_pnl', 'total_pnl'):
                return f"${value:+,.2f}" if value is not None else ""
            elif col_key == 'ud_vol_ratio':
                return f"{value:.2f}" if value is not None else ""
            elif col_key in ('watch_date', 'breakout_date', 'entry_date', 'earnings_date', 'close_date', 'watching_exited_since'):
                if isinstance(value, (date, datetime)):
                    return value.strftime('%Y-%m-%d')
                return str(value) if value else ""
            elif col_key == 'notes':
                # Truncate long notes for display
                if value:
                    return str(value)[:50] + "..." if len(str(value)) > 50 else str(value)
                return ""
            elif value is None:
                return ""
            return str(value)
        
        elif role == Qt.ItemDataRole.BackgroundRole:
            state = row_data.get('state', 0)
            if state == -2:
                return QBrush(QColor(80, 45, 45))       # Red - stopped out
            elif state == -1.5:
                return QBrush(QColor(75, 45, 85))       # Purple - watching exited
            elif state == -1:
                return QBrush(QColor(85, 65, 40))       # Amber - failed setup
            elif state == 0:
                return QBrush(QColor(45, 55, 75))       # Blue - watching
            elif state >= 1:
                return QBrush(QColor(45, 75, 55))       # Green - in position
        
        elif role == Qt.ItemDataRole.ForegroundRole:
            # Color all P&L and distance columns
            if col_key in ('dist_from_pivot', 'unrealized_pct', 'unrealized_pnl',
                          'realized_pnl', 'realized_pnl_pct', 'total_pnl') and value is not None:
                if value > 0:
                    return QBrush(QColor(80, 220, 100))  # Green
                elif value < 0:
                    return QBrush(QColor(255, 100, 100))  # Red
        
        elif role == Qt.ItemDataRole.TextAlignmentRole:
            if col_key in ('pivot', 'last_price', 'avg_cost', 'stop_price',
                          'dist_from_pivot', 'unrealized_pct', 'unrealized_pnl',
                          'realized_pnl', 'realized_pnl_pct', 'total_pnl',
                          'rs_rating', 'rs_3mo', 'rs_6mo', 'eps_rating', 'comp_rating',
                          'industry_rank', 'base_depth', 'base_length', 'fund_count',
                          'funds_qtr_chg', 'entry_score', 'total_shares',
                          'ud_vol_ratio', 'prior_uptrend', 'hard_stop_pct', 'health_score',
                          'close_price', 'original_pivot', 'ma_test_count'):
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        
        elif role == Qt.ItemDataRole.UserRole:
            return value
        
        return None
    
    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return self._column_names[section]
            else:
                return section + 1
        return None
    
    def get_position_data(self, row: int) -> Optional[Dict[str, Any]]:
        if 0 <= row < len(self._data):
            return self._data[row]
        return None
    
    def get_column_key(self, column: int) -> str:
        return self._column_keys[column]


class AdvancedFilterProxyModel(QSortFilterProxyModel):
    """Proxy model supporting text multi-select and numeric comparison filters."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._filters: Dict[int, Any] = {}
        self._state_filter: Optional[float] = None  # float to support -1.5

    def set_column_filter(self, column: int, filter_value):
        """Set filter for a column. None to clear."""
        if filter_value is None:
            if column in self._filters:
                del self._filters[column]
        else:
            self._filters[column] = filter_value
        self.invalidateFilter()

    def set_state_filter(self, state: Optional[float]):
        self._state_filter = state
        self.invalidateFilter()
    
    def clear_filters(self):
        self._filters.clear()
        self._state_filter = None
        self.invalidateFilter()
    
    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        model = self.sourceModel()
        
        # Check state filter
        if self._state_filter is not None:
            state_col = model._column_keys.index('state')
            state_index = model.index(source_row, state_col, source_parent)
            state_value = model.data(state_index, Qt.ItemDataRole.UserRole)
            if state_value != self._state_filter:
                return False
        
        # Check column filters
        for column, filter_value in self._filters.items():
            index = model.index(source_row, column, source_parent)
            raw_value = model.data(index, Qt.ItemDataRole.UserRole)
            display_value = model.data(index, Qt.ItemDataRole.DisplayRole)
            
            if isinstance(filter_value, set):
                # Text multi-select filter
                if display_value not in filter_value and str(raw_value) not in filter_value:
                    return False
            
            elif isinstance(filter_value, dict):
                op = filter_value.get('operator')
                filter_type = filter_value.get('type', 'numeric')

                if filter_type == 'date':
                    # Date comparison filter
                    if raw_value is None:
                        return False

                    # Convert to date if needed
                    if isinstance(raw_value, datetime):
                        date_value = raw_value.date()
                    elif isinstance(raw_value, date):
                        date_value = raw_value
                    else:
                        return False

                    target = filter_value.get('value')
                    target2 = filter_value.get('value2')

                    if op == '=' and date_value != target:
                        return False
                    elif op == '>' and date_value <= target:
                        return False
                    elif op == '≥' and date_value < target:
                        return False
                    elif op == '<' and date_value >= target:
                        return False
                    elif op == '≤' and date_value > target:
                        return False
                    elif op == 'between' and target2:
                        min_date, max_date = min(target, target2), max(target, target2)
                        if date_value < min_date or date_value > max_date:
                            return False
                else:
                    # Numeric comparison filter
                    if raw_value is None:
                        return False

                    try:
                        num_value = float(raw_value)
                    except (ValueError, TypeError):
                        return False

                    target = filter_value.get('value', 0)
                    target2 = filter_value.get('value2', 0)

                    if op == '=' and num_value != target:
                        return False
                    elif op == '≠' and num_value == target:
                        return False
                    elif op == '>' and num_value <= target:
                        return False
                    elif op == '≥' and num_value < target:
                        return False
                    elif op == '<' and num_value >= target:
                        return False
                    elif op == '≤' and num_value > target:
                        return False
                    elif op == 'between':
                        min_val, max_val = min(target, target2), max(target, target2)
                        if num_value < min_val or num_value > max_val:
                            return False

        return True


class PositionTableView(QDialog):
    """Spreadsheet-style view with Excel-like filtering and customizable columns."""

    position_edited = pyqtSignal(int, dict)

    # Default config path
    DEFAULT_CONFIG_PATH = "c:/Trading/canslim_monitor/user_config.yaml"

    # Default column widths for reference
    DEFAULT_COLUMN_WIDTHS = [
        70, 60, 85, 100, 70, 75, 55, 60, 75, 70, 75,  # Symbol through Total $
        40, 50, 50, 40, 50, 45, 45, 60, 50, 60, 50,   # RS through Length
        55, 60, 55, 65, 50, 50, 85, 85, 85, 85,       # U/D Vol through Earnings
        55, 75, 55, 70, 50,                            # Shares through Health
        70, 85, 90, 55,                                # Close$, CloseDate, CloseReason, Real%
        75, 60, 85,                                    # OrigPivot, MATsts, WatchExit (State -1.5)
        120                                            # Notes
    ]

    def __init__(self, db_session, parent=None, config_path: str = None):
        super().__init__(parent)
        self.db_session = db_session
        self._position_id_map: Dict[int, int] = {}
        self._filter_buttons: Dict[int, ColumnFilterButton] = {}
        self._config_path = config_path or self.DEFAULT_CONFIG_PATH

        # View state: tracks hidden columns, column order, widths, and sorting
        self._hidden_columns: Set[int] = set()
        self._column_order: List[int] = list(range(len(PositionTableModel.COLUMNS)))
        self._column_widths: Dict[int, int] = {}  # logical_index -> width
        self._sort_columns: List[tuple] = []  # [(column_index, ascending), ...]

        # Multiple views support
        self._current_view_name: str = "Default"
        self._saved_views: Dict[str, dict] = {}  # view_name -> view_state
        self._loading_view: bool = False  # Prevent save during load

        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint
        )

        self._setup_ui()
        self._load_view_state()  # Load saved column visibility and order
        self._apply_view_state()  # Apply to table
        self._load_data()
    
    def _setup_ui(self):
        self.setWindowTitle("Position Database View")
        self.setMinimumSize(1200, 700)
        self.resize(1400, 800)
        
        # Overall window style
        self.setStyleSheet("""
            QDialog {
                background-color: #2D2D30;
            }
            QLabel {
                color: #E0E0E0;
            }
            QComboBox {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                border-radius: 3px;
                padding: 4px;
                color: #E0E0E0;
            }
            QPushButton {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                border-radius: 3px;
                padding: 6px 12px;
                color: #E0E0E0;
            }
            QPushButton:hover {
                background-color: #4A4A4D;
                border: 1px solid #0078D4;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("All Positions")
        title_font = QFont()
        title_font.setBold(True)
        title_font.setPointSize(14)
        title.setFont(title_font)
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        self.count_label = QLabel("0 positions")
        self.count_label.setStyleSheet("color: #AAAAAA; font-style: italic;")
        header_layout.addWidget(self.count_label)
        
        layout.addLayout(header_layout)

        # View management row
        view_layout = QHBoxLayout()
        view_layout.setSpacing(8)

        view_label = QLabel("View:")
        view_label.setStyleSheet("color: #AAAAAA;")
        view_layout.addWidget(view_label)

        self.view_combo = QComboBox()
        self.view_combo.setMinimumWidth(150)
        self.view_combo.currentTextChanged.connect(self._on_view_changed)
        view_layout.addWidget(self.view_combo)

        save_view_btn = QPushButton("Save")
        save_view_btn.setToolTip("Save changes to current view")
        save_view_btn.clicked.connect(self._save_current_view)
        view_layout.addWidget(save_view_btn)

        save_as_btn = QPushButton("Save As...")
        save_as_btn.setToolTip("Save as a new named view")
        save_as_btn.clicked.connect(self._save_view_as)
        view_layout.addWidget(save_as_btn)

        delete_view_btn = QPushButton("Delete")
        delete_view_btn.setToolTip("Delete current view")
        delete_view_btn.clicked.connect(self._delete_current_view)
        view_layout.addWidget(delete_view_btn)

        view_layout.addSpacing(20)

        # Export button with menu
        export_btn = QPushButton("Export...")
        export_btn.setToolTip("Export visible data to file")
        export_menu = QMenu(export_btn)
        export_menu.setStyleSheet("""
            QMenu {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
            }
            QMenu::item {
                color: #E0E0E0;
                padding: 6px 20px;
            }
            QMenu::item:selected {
                background-color: #0078D4;
            }
        """)
        export_csv_action = export_menu.addAction("Export to CSV")
        export_csv_action.triggered.connect(self._export_to_csv)
        export_excel_action = export_menu.addAction("Export to Excel")
        export_excel_action.triggered.connect(self._export_to_excel)
        export_btn.setMenu(export_menu)
        view_layout.addWidget(export_btn)

        view_layout.addStretch()
        layout.addLayout(view_layout)

        # Quick filters row
        quick_filter_layout = QHBoxLayout()
        
        # State filter dropdown
        state_layout = QFormLayout()
        self.state_filter_combo = QComboBox()
        self.state_filter_combo.addItem("All States", None)
        self.state_filter_combo.addItem("📋 Watching (0)", 0)
        self.state_filter_combo.addItem("🎯 Entry Ready (1)", 1)
        self.state_filter_combo.addItem("💰 In Position (2)", 2)
        self.state_filter_combo.addItem("📈 Pyramiding (3)", 3)
        self.state_filter_combo.addItem("⚠️ Failed Setup (-1)", -1)
        self.state_filter_combo.addItem("👁️ Exited Watch (-1.5)", -1.5)
        self.state_filter_combo.addItem("🛑 Stopped Out (-2)", -2)
        self.state_filter_combo.currentIndexChanged.connect(self._on_state_filter_changed)
        state_layout.addRow("State:", self.state_filter_combo)
        quick_filter_layout.addLayout(state_layout)
        
        quick_filter_layout.addStretch()
        
        # Info label
        info_label = QLabel("💡 Click ▼ to filter | Right-click headers to show/hide columns | Drag to reorder")
        info_label.setStyleSheet("color: #888888; font-style: italic;")
        quick_filter_layout.addWidget(info_label)
        
        clear_btn = QPushButton("Clear All Filters")
        clear_btn.clicked.connect(self._clear_filters)
        quick_filter_layout.addWidget(clear_btn)
        
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self._load_data)
        quick_filter_layout.addWidget(refresh_btn)
        
        layout.addLayout(quick_filter_layout)
        
        # Table model and proxy
        self.table_model = PositionTableModel()
        self.proxy_model = AdvancedFilterProxyModel()
        self.proxy_model.setSourceModel(self.table_model)
        self.proxy_model.setSortCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        
        # Filter buttons row
        self.filter_row = QWidget()
        self.filter_row.setFixedHeight(24)
        self.filter_row.setStyleSheet("background-color: #3C3C3F;")
        filter_row_layout = QHBoxLayout(self.filter_row)
        filter_row_layout.setContentsMargins(0, 0, 0, 0)
        filter_row_layout.setSpacing(0)
        
        for i, (key, name) in enumerate(PositionTableModel.COLUMNS):
            btn = ColumnFilterButton(i, key, name)
            btn.setFixedHeight(22)
            btn.filter_changed.connect(self._on_column_filter_changed)
            self._filter_buttons[i] = btn
            filter_row_layout.addWidget(btn)
        
        filter_row_layout.addStretch()
        layout.addWidget(self.filter_row)
        
        # Table view
        self.table_view = QTableView()
        self.table_view.setModel(self.proxy_model)
        self.table_view.setSortingEnabled(True)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.verticalHeader().setVisible(False)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # Enable column reordering via drag-and-drop
        self.table_view.horizontalHeader().setSectionsMovable(True)
        self.table_view.horizontalHeader().sectionMoved.connect(self._on_column_moved)

        # Add context menu to header for column visibility
        self.table_view.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.horizontalHeader().customContextMenuRequested.connect(self._on_header_context_menu)

        # Track column width changes for persistence
        self.table_view.horizontalHeader().sectionResized.connect(self._on_column_resized)

        # Track sort changes for persistence
        self.table_view.horizontalHeader().sortIndicatorChanged.connect(self._on_sort_changed)

        # Double-click to edit
        self.table_view.doubleClicked.connect(self._on_double_click)

        # Right-click context menu
        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self._on_table_context_menu)

        # Set column widths from defaults
        header = self.table_view.horizontalHeader()
        for i, width in enumerate(self.DEFAULT_COLUMN_WIDTHS):
            if i < len(PositionTableModel.COLUMNS):
                header.resizeSection(i, width)
                self._column_widths[i] = width
        
        # Sync filter button widths with columns
        self.table_view.horizontalHeader().sectionResized.connect(self._sync_filter_widths)
        QTimer.singleShot(100, self._sync_filter_widths)
        
        # Style
        self.table_view.setStyleSheet("""
            QTableView {
                background-color: #2D2D30;
                alternate-background-color: #363638;
                gridline-color: #4A4A4D;
                selection-background-color: #0078D4;
                color: #E0E0E0;
            }
            QTableView::item {
                padding: 4px;
                border-bottom: 1px solid #4A4A4D;
            }
            QTableView::item:selected {
                background-color: #0078D4;
                color: white;
            }
            QHeaderView::section {
                background-color: #3C3C3F;
                color: #FFFFFF;
                padding: 6px 4px;
                border: none;
                border-right: 1px solid #4A4A4D;
                border-bottom: 1px solid #4A4A4D;
                font-weight: bold;
                font-size: 11px;
            }
            QHeaderView::section:hover {
                background-color: #4A4A4D;
            }
        """)
        
        layout.addWidget(self.table_view)
        
        # Bottom buttons
        button_layout = QHBoxLayout()
        
        export_btn = QPushButton("📊 Export CSV")
        export_btn.clicked.connect(self._export_csv)
        button_layout.addWidget(export_btn)
        
        button_layout.addStretch()
        
        edit_btn = QPushButton("✏️ Edit Selected")
        edit_btn.clicked.connect(self._edit_selected)
        button_layout.addWidget(edit_btn)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.close)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
    
    def _sync_filter_widths(self):
        """Sync filter button widths and positions with table column widths and visual order."""
        header = self.table_view.horizontalHeader()

        # Get the filter row layout
        layout = self.filter_row.layout()
        if not layout:
            return

        # Build visual order: list of logical indices in visual order
        visual_order = []
        for visual_index in range(header.count()):
            logical_index = header.logicalIndex(visual_index)
            visual_order.append(logical_index)

        # Remove all widgets from layout (but keep references)
        while layout.count():
            item = layout.takeAt(0)
            # Don't delete the widget, just remove from layout

        # Re-add buttons in visual order with correct widths
        for visual_index, logical_index in enumerate(visual_order):
            btn = self._filter_buttons.get(logical_index)
            if btn:
                is_hidden = header.isSectionHidden(logical_index)
                btn.setVisible(not is_hidden)
                btn.setFixedWidth(header.sectionSize(logical_index))
                layout.addWidget(btn)

        # Add stretch at end
        layout.addStretch()
    
    def _load_data(self):
        """Load all positions from database."""
        try:
            from canslim_monitor.data.models import Position
            
            positions = self.db_session.query(Position).all()
            
            position_data = []
            self._position_id_map.clear()
            
            for i, pos in enumerate(positions):
                # Compute distance from pivot: ((last_price - pivot) / pivot) * 100
                dist_from_pivot = None
                if pos.last_price and pos.pivot and pos.pivot > 0:
                    dist_from_pivot = ((pos.last_price - pos.pivot) / pos.pivot) * 100
                
                # Compute unrealized P&L % : ((last_price - avg_cost) / avg_cost) * 100
                unrealized_pct = None
                if pos.last_price and pos.avg_cost and pos.avg_cost > 0:
                    unrealized_pct = ((pos.last_price - pos.avg_cost) / pos.avg_cost) * 100
                
                # Compute unrealized P&L $: (last_price - avg_cost) * total_shares
                unrealized_pnl = None
                if pos.last_price and pos.avg_cost and pos.total_shares:
                    unrealized_pnl = (pos.last_price - pos.avg_cost) * pos.total_shares
                
                # Get realized P&L if it exists in the model
                realized_pnl = getattr(pos, 'realized_pnl', None)
                
                # Compute total P&L: realized + unrealized
                total_pnl = None
                if realized_pnl is not None or unrealized_pnl is not None:
                    total_pnl = (realized_pnl or 0) + (unrealized_pnl or 0)
                
                data = {
                    'id': pos.id,
                    'symbol': pos.symbol,
                    'portfolio': pos.portfolio,
                    'state': pos.state,
                    'pattern': pos.pattern,
                    'pivot': pos.pivot,
                    'last_price': pos.last_price,
                    'dist_from_pivot': dist_from_pivot,    # Computed
                    'unrealized_pct': unrealized_pct,      # Computed
                    'unrealized_pnl': unrealized_pnl,      # Computed
                    'realized_pnl': realized_pnl,          # From DB
                    'realized_pnl_pct': getattr(pos, 'realized_pnl_pct', None),
                    'total_pnl': total_pnl,                # Computed
                    'rs_rating': pos.rs_rating,
                    'rs_3mo': pos.rs_3mo,
                    'rs_6mo': pos.rs_6mo,
                    'eps_rating': pos.eps_rating,
                    'comp_rating': pos.comp_rating,
                    'smr_rating': pos.smr_rating,
                    'ad_rating': pos.ad_rating,
                    'industry_rank': pos.industry_rank,
                    'industry_eps_rank': pos.industry_eps_rank,
                    'industry_rs_rank': pos.industry_rs_rank,
                    'base_stage': pos.base_stage,
                    'base_depth': pos.base_depth,
                    'base_length': pos.base_length,
                    'ud_vol_ratio': pos.ud_vol_ratio,
                    'prior_uptrend': pos.prior_uptrend,
                    'fund_count': pos.fund_count,
                    'funds_qtr_chg': pos.funds_qtr_chg,
                    'breakout_vol_pct': pos.breakout_vol_pct,
                    'breakout_price_pct': pos.breakout_price_pct,
                    'entry_grade': pos.entry_grade,
                    'entry_score': pos.entry_score,
                    'watch_date': pos.watch_date,
                    'entry_date': pos.entry_date,
                    'breakout_date': pos.breakout_date,
                    'earnings_date': pos.earnings_date,
                    'total_shares': pos.total_shares,
                    'avg_cost': pos.avg_cost,
                    'hard_stop_pct': pos.hard_stop_pct,
                    'stop_price': pos.stop_price,
                    'health_score': pos.health_score,
                    'notes': pos.notes,
                    # Entry tranches
                    'e1_shares': pos.e1_shares,
                    'e1_price': pos.e1_price,
                    'e2_shares': pos.e2_shares,
                    'e2_price': pos.e2_price,
                    'e3_shares': pos.e3_shares,
                    'e3_price': pos.e3_price,
                    # Exit / Close
                    'tp1_sold': pos.tp1_sold,
                    'tp1_price': pos.tp1_price,
                    'tp1_date': pos.tp1_date,
                    'tp2_sold': pos.tp2_sold,
                    'tp2_price': pos.tp2_price,
                    'tp2_date': pos.tp2_date,
                    'close_price': getattr(pos, 'close_price', None),
                    'close_date': getattr(pos, 'close_date', None),
                    'close_reason': getattr(pos, 'close_reason', None),
                    # State -1.5 (WATCHING_EXITED) fields
                    'original_pivot': getattr(pos, 'original_pivot', None),
                    'ma_test_count': getattr(pos, 'ma_test_count', None),
                    'watching_exited_since': getattr(pos, 'watching_exited_since', None),
                }
                position_data.append(data)
                self._position_id_map[i] = pos.id
            
            self.table_model.load_positions(position_data)
            
            # Update filter buttons with unique values
            for i, btn in self._filter_buttons.items():
                if COLUMN_TYPES.get(btn.column_key) not in ('int', 'float'):
                    btn.set_unique_values(self.table_model.get_unique_values(i))
            
            self._update_count()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load positions: {e}")
    
    def _update_count(self):
        visible = self.proxy_model.rowCount()
        total = self.table_model.rowCount()
        if visible == total:
            self.count_label.setText(f"{total} positions")
        else:
            self.count_label.setText(f"{visible} of {total} positions")
    
    def _on_state_filter_changed(self, index: int):
        state = self.state_filter_combo.currentData()
        self.proxy_model.set_state_filter(state)
        self._update_count()
    
    def _on_column_filter_changed(self, column: int, filter_value):
        self.proxy_model.set_column_filter(column, filter_value)
        self._update_count()
    
    def _clear_filters(self):
        self.state_filter_combo.setCurrentIndex(0)
        for btn in self._filter_buttons.values():
            btn.clear_filter()
        self.proxy_model.clear_filters()
        self._update_count()
    
    def _on_double_click(self, index: QModelIndex):
        source_index = self.proxy_model.mapToSource(index)
        row = source_index.row()
        position_data = self.table_model.get_position_data(row)
        if position_data:
            self._open_edit_dialog(position_data)
    
    def _edit_selected(self):
        indexes = self.table_view.selectionModel().selectedRows()
        if not indexes:
            QMessageBox.information(self, "No Selection", "Please select a position to edit.")
            return
        
        source_index = self.proxy_model.mapToSource(indexes[0])
        row = source_index.row()
        position_data = self.table_model.get_position_data(row)
        if position_data:
            self._open_edit_dialog(position_data)
    
    def _open_edit_dialog(self, position_data: Dict[str, Any]):
        from canslim_monitor.gui.transition_dialogs import EditPositionDialog
        import logging
        logger = logging.getLogger('canslim.position_table')

        dialog = EditPositionDialog(position_data, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            updated_data = dialog.get_result()
            position_id = position_data.get('id')

            # Log key exit/close fields
            logger.info(f"=== TABLE VIEW EDIT SAVE: {position_data.get('symbol')} (id={position_id}) ===")
            close_fields = ['close_price', 'close_date', 'close_reason', 'realized_pnl', 'realized_pnl_pct',
                           'tp1_sold', 'tp1_price', 'tp2_sold', 'tp2_price']
            for key in close_fields:
                if key in updated_data and updated_data[key] is not None:
                    logger.info(f"  {key}: {updated_data[key]}")

            if position_id:
                self.position_edited.emit(position_id, updated_data)
                self._load_data()
    
    def _export_csv(self):
        from PyQt6.QtWidgets import QFileDialog
        import csv
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Positions", "positions.csv", "CSV Files (*.csv)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                headers = [col[1] for col in PositionTableModel.COLUMNS]
                writer.writerow(headers)
                
                for row in range(self.proxy_model.rowCount()):
                    row_data = []
                    for col in range(self.proxy_model.columnCount()):
                        index = self.proxy_model.index(row, col)
                        value = self.proxy_model.data(index, Qt.ItemDataRole.DisplayRole)
                        row_data.append(value if value else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {self.proxy_model.rowCount()} positions to:\n{filename}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")

    def _on_table_context_menu(self, pos: QPoint):
        """Handle right-click context menu on table rows."""
        # Get the index at the click position
        index = self.table_view.indexAt(pos)
        if not index.isValid():
            return

        # Get the source row and position data
        source_index = self.proxy_model.mapToSource(index)
        row = source_index.row()
        position_data = self.table_model.get_position_data(row)

        if not position_data:
            return

        position_id = position_data.get('id')
        state = position_data.get('state', 0)
        symbol = position_data.get('symbol', 'Unknown')

        menu = QMenu(self)

        # Edit action
        edit_action = menu.addAction("✏️ Edit Position")
        edit_action.triggered.connect(lambda: self._open_edit_dialog(position_data))

        # Score details action
        score_action = menu.addAction("📊 View Score Details")
        score_action.triggered.connect(lambda: self._show_score_details(position_id))

        # View alerts action
        alerts_action = menu.addAction("🔔 View Alerts")
        alerts_action.triggered.connect(lambda: self._show_position_alerts(symbol, position_id))

        menu.addSeparator()

        # State transition actions
        from canslim_monitor.gui.state_config import get_valid_transitions
        valid_transitions = get_valid_transitions(state)

        if valid_transitions:
            move_menu = menu.addMenu("📦 Move to...")
            for transition in valid_transitions:
                to_state = transition.to_state
                to_name = STATES[to_state].display_name
                action = move_menu.addAction(f"{to_name}")
                action.triggered.connect(
                    lambda checked, ts=to_state, pid=position_id, cs=state: self._trigger_transition(pid, cs, ts)
                )

        menu.addSeparator()

        # Quick actions based on state
        if state >= 1:  # In position
            stop_action = menu.addAction("🛑 Stop Out")
            stop_action.triggered.connect(
                lambda: self._trigger_transition(position_id, state, -2)
            )

            close_action = menu.addAction("✅ Close Position")
            close_action.triggered.connect(
                lambda: self._trigger_transition(position_id, state, -1)
            )

        menu.addSeparator()

        # Delete action
        delete_action = menu.addAction("🗑️ Delete")
        delete_action.triggered.connect(lambda: self._delete_position(position_id, symbol))

        # Show menu at cursor position
        menu.exec(self.table_view.viewport().mapToGlobal(pos))

    def _show_score_details(self, position_id: int):
        """Show detailed score breakdown for a position."""
        import json
        from canslim_monitor.utils.scoring import CANSLIMScorer
        from canslim_monitor.data.models import Position
        from PyQt6.QtWidgets import QTextEdit

        try:
            position = self.db_session.query(Position).filter_by(id=position_id).first()

            if not position:
                QMessageBox.warning(self, "Error", "Position not found")
                return

            # Get current market regime (default to BULLISH)
            market_regime = 'BULLISH'
            try:
                from canslim_monitor.data.models import MarketRegime
                market_config = self.db_session.query(MarketRegime).order_by(MarketRegime.id.desc()).first()
                if market_config:
                    market_regime = market_config.regime
            except:
                pass

            scorer = CANSLIMScorer()

            # Try to load stored details first
            details = None
            if position.entry_score_details:
                try:
                    details = json.loads(position.entry_score_details)
                    if 'total_score' not in details or 'grade' not in details:
                        details = None
                except (json.JSONDecodeError, TypeError):
                    details = None

            # If no stored details, calculate static score
            if details is None:
                position_data = {
                    'pattern': position.pattern,
                    'rs_rating': position.rs_rating,
                    'eps_rating': position.eps_rating,
                    'ad_rating': position.ad_rating,
                    'base_stage': position.base_stage,
                    'base_depth': position.base_depth,
                    'base_length': position.base_length,
                    'ud_vol_ratio': position.ud_vol_ratio,
                    'fund_count': position.fund_count,
                    'group_rank': position.group_rank,
                    'comp_rating': position.comp_rating,
                    'prior_uptrend': position.prior_uptrend,
                }
                score, grade, details = scorer.calculate_score(position_data, market_regime)
            else:
                score = details.get('total_score', position.entry_score or 0)
                grade = details.get('grade', position.entry_grade or 'F')

            # Format details text
            details_text = scorer.format_details_text(details, symbol=position.symbol, pivot=position.pivot)

            # Fetch fresh earnings date from API and update database
            earnings_date = self._fetch_and_update_earnings_date(position)
            # Fall back to database value if API didn't find anything
            if earnings_date is None:
                earnings_date = position.earnings_date

            # Show in a dialog
            from datetime import date as date_type

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Score Details: {position.symbol}")
            dialog.setMinimumSize(720, 750)

            layout = QVBoxLayout(dialog)

            # Header with symbol and grade
            header = QLabel(f"<h2>{position.symbol} - Grade: {grade} (Score: {score})</h2>")
            header.setStyleSheet("color: #E0E0E0;")
            layout.addWidget(header)

            # Earnings warning section (uses freshly fetched or database earnings_date)
            earnings_warning = self._create_earnings_warning_label(position.symbol, earnings_date)
            if earnings_warning:
                layout.addWidget(earnings_warning)

            # Score details in monospace text
            text_edit = QTextEdit()
            text_edit.setReadOnly(True)
            text_edit.setFontFamily("Consolas, Monaco, monospace")
            text_edit.setPlainText(details_text)
            text_edit.setStyleSheet("""
                QTextEdit {
                    background-color: #1E1E1E;
                    color: #D4D4D4;
                    font-size: 13px;
                    padding: 10px;
                    line-height: 1.4;
                }
            """)
            layout.addWidget(text_edit)

            # Buttons
            btn_layout = QHBoxLayout()

            # Recalculate button - call parent's recalculate method if available
            recalc_btn = QPushButton("🔄 Recalculate")
            recalc_btn.setToolTip("Fetch latest data and recalculate score with execution feasibility")

            def do_recalculate():
                # Call parent's recalculate method if available (kanban_window)
                if self.parent() and hasattr(self.parent(), '_recalculate_score'):
                    # Pass dialog as parent so it can be closed and score details reopened after completion
                    self.parent()._recalculate_score(position_id, dialog)
                else:
                    QMessageBox.information(
                        dialog, "Info",
                        "Recalculate is only available from the main window.\n"
                        "Please use the kanban view for full recalculation."
                    )

            recalc_btn.clicked.connect(do_recalculate)
            btn_layout.addWidget(recalc_btn)

            btn_layout.addStretch()

            close_btn = QPushButton("Close")
            close_btn.clicked.connect(dialog.accept)
            btn_layout.addWidget(close_btn)

            layout.addLayout(btn_layout)

            dialog.exec()

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to show score details: {e}")

    def _create_earnings_warning_label(self, symbol: str, earnings_date) -> QLabel:
        """
        Create earnings warning label with progressive color coding.

        Args:
            symbol: Stock symbol
            earnings_date: Earnings date (date object or None)

        Returns:
            QLabel with appropriate styling, or None if no warning needed
        """
        from datetime import date as date_type

        # Get config from parent if available
        config = {}
        if self.parent() and hasattr(self.parent(), 'config'):
            config = self.parent().config

        earnings_config = config.get('earnings', {})
        thresholds = earnings_config.get('warning_thresholds', {})
        critical_days = thresholds.get('critical', 5)
        caution_days = thresholds.get('caution', 10)

        label = QLabel()
        label.setWordWrap(True)

        if earnings_date:
            today = date_type.today()
            days_until = (earnings_date - today).days

            if days_until < 0:
                label.setText(
                    f"⚠️ Earnings date ({earnings_date}) appears to be in the past."
                )
                label.setStyleSheet("""
                    QLabel {
                        color: #856404;
                        background-color: #FFF3CD;
                        padding: 8px;
                        border-radius: 4px;
                        font-weight: bold;
                    }
                """)

            elif days_until <= critical_days:
                label.setText(
                    f"⚠️ EARNINGS IN {days_until} DAY{'S' if days_until != 1 else ''} "
                    f"({earnings_date.strftime('%b %d')}) - HIGH RISK!"
                )
                label.setStyleSheet("""
                    QLabel {
                        color: white;
                        background-color: #DC3545;
                        padding: 10px;
                        border-radius: 4px;
                        font-weight: bold;
                        font-size: 13px;
                    }
                """)

            elif days_until <= caution_days:
                label.setText(
                    f"⚠️ Earnings in {days_until} days ({earnings_date.strftime('%b %d')}) - "
                    f"Monitor closely."
                )
                label.setStyleSheet("""
                    QLabel {
                        color: #856404;
                        background-color: #FFC107;
                        padding: 8px;
                        border-radius: 4px;
                        font-weight: bold;
                    }
                """)

            else:
                label.setText(
                    f"📅 Next earnings: {earnings_date.strftime('%b %d, %Y')} ({days_until} days)"
                )
                label.setStyleSheet("""
                    QLabel {
                        color: #155724;
                        background-color: #D4EDDA;
                        padding: 6px;
                        border-radius: 4px;
                    }
                """)

            return label

        else:
            label.setText(
                f"⚠️ No earnings date for {symbol}. Check MarketSurge or enter manually."
            )
            label.setStyleSheet("""
                QLabel {
                    color: #856404;
                    background-color: #FFF3CD;
                    padding: 8px;
                    border-radius: 4px;
                    font-style: italic;
                }
            """)
            return label

    def _fetch_and_update_earnings_date(self, position):
        """
        Fetch earnings date from API and update database if found.

        Args:
            position: Position model instance

        Returns:
            Earnings date if found, None otherwise
        """
        try:
            from canslim_monitor.integrations.polygon_client import PolygonClient

            # Get config from parent if available
            config = {}
            if self.parent() and hasattr(self.parent(), 'config'):
                config = self.parent().config

            market_data_config = config.get('market_data', {})
            polygon_config = config.get('polygon', {})

            api_key = market_data_config.get('api_key') or polygon_config.get('api_key', '')
            base_url = market_data_config.get('base_url') or polygon_config.get('base_url', 'https://api.polygon.io')

            if not api_key:
                return None

            # Fetch earnings date (tries Polygon, then Yahoo Finance)
            polygon_client = PolygonClient(api_key=api_key, base_url=base_url)
            earnings_date = polygon_client.get_next_earnings_date(position.symbol)

            if earnings_date:
                # Update position in database
                if position.earnings_date != earnings_date:
                    position.earnings_date = earnings_date
                    self.db_session.commit()

                return earnings_date

        except Exception as e:
            import logging
            logging.getLogger('canslim.gui').warning(f"Error fetching earnings date: {e}")

        return None

    def _show_position_alerts(self, symbol: str, position_id: int):
        """Show alerts dialog for position using the AlertService."""
        try:
            from canslim_monitor.gui.alerts import PositionAlertDialog
            from canslim_monitor.services.alert_service import AlertService
            import logging

            # Get db_session_factory from parent (kanban_window)
            db_session_factory = None
            if self.parent() and hasattr(self.parent(), 'db'):
                db_session_factory = self.parent().db.get_new_session

            # Create alert service
            alert_service = AlertService(
                db_session_factory=db_session_factory,
                logger=logging.getLogger('canslim.alerts')
            )

            # Fetch alerts for this symbol using the proper service
            alerts = alert_service.get_recent_alerts(
                symbol=symbol,
                hours=24 * 90,  # 90 days
                limit=500
            )

            if not alerts:
                QMessageBox.information(self, "No Alerts", f"No alerts found for {symbol}")
                return

            # Show dialog with alert service for acknowledgment support
            dialog = PositionAlertDialog(
                symbol=symbol,
                alerts=alerts,
                parent=self,
                alert_service=alert_service,
                db_session_factory=db_session_factory
            )
            dialog.alert_acknowledged.connect(lambda _: self._load_data())
            dialog.show()  # Modeless like kanban_window

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to show alerts: {e}")

    def _trigger_transition(self, position_id: int, from_state: int, to_state: int):
        """Trigger a state transition for the position."""
        try:
            from canslim_monitor.data.models import Position

            position = self.db_session.query(Position).filter_by(id=position_id).first()
            if not position:
                QMessageBox.warning(self, "Error", "Position not found")
                return

            # Check if transition requires a dialog (has required or optional fields)
            from canslim_monitor.gui.state_config import get_transition
            transition = get_transition(from_state, to_state)

            if transition and (transition.required_fields or transition.optional_fields):
                # Use TransitionDialog for all transitions needing input
                from canslim_monitor.gui.transition_dialogs import TransitionDialog

                # Prepare current data for dialog
                current_data = {
                    'symbol': position.symbol,
                    'pivot': position.pivot,
                    'stop_price': position.stop_price,
                    'avg_cost': position.avg_cost,
                    'total_shares': position.total_shares,
                    'last_price': position.last_price,
                    'e1_shares': position.e1_shares,
                    'e1_price': position.e1_price,
                    'e2_shares': position.e2_shares,
                    'e2_price': position.e2_price,
                    'e3_shares': position.e3_shares,
                    'e3_price': position.e3_price,
                    'entry_date': position.entry_date,
                    'breakout_date': position.breakout_date,
                    'watch_date': position.watch_date,
                    'current_pnl_pct': position.current_pnl_pct,
                    'notes': position.notes,
                }

                dialog = TransitionDialog(
                    symbol=position.symbol,
                    from_state=from_state,
                    to_state=to_state,
                    current_data=current_data,
                    parent=self
                )

                if dialog.exec() == QDialog.DialogCode.Accepted:
                    result = dialog.get_result()

                    # Map exit_* fields to close_* fields for Position model compatibility
                    if to_state < 0:
                        if 'exit_price' in result:
                            result['close_price'] = result['exit_price']
                        if 'exit_date' in result:
                            result['close_date'] = result['exit_date']
                        if 'exit_reason' in result:
                            result['close_reason'] = result['exit_reason']

                        # Calculate realized P&L
                        close_price = result.get('close_price') or result.get('exit_price', 0)
                        avg_cost = position.avg_cost or 0
                        if close_price and avg_cost > 0:
                            result['realized_pnl_pct'] = ((close_price - avg_cost) / avg_cost) * 100
                            total_shares = position.total_shares or 0
                            if total_shares > 0:
                                result['realized_pnl'] = (close_price - avg_cost) * total_shares

                    # Apply the transition result
                    for key, value in result.items():
                        if hasattr(position, key):
                            setattr(position, key, value)
                    position.state = to_state
                    self.db_session.commit()
                    self._load_data()

                    # Emit signal so parent window updates
                    self.position_edited.emit(position_id, result)
                return

            # Simple state change (no dialog required)
            position.state = to_state
            self.db_session.commit()
            self._load_data()

            # Emit signal so parent window updates
            self.position_edited.emit(position_id, {'state': to_state})

        except Exception as e:
            self.db_session.rollback()
            QMessageBox.warning(self, "Error", f"Failed to change state: {e}")

    def _delete_position(self, position_id: int, symbol: str):
        """Delete a position after confirmation."""
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete position '{symbol}'?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                from canslim_monitor.data.models import Position

                position = self.db_session.query(Position).filter_by(id=position_id).first()
                if position:
                    self.db_session.delete(position)
                    self.db_session.commit()
                    self._load_data()
                    QMessageBox.information(self, "Deleted", f"Position '{symbol}' has been deleted.")

                    # Emit signal so parent window updates
                    self.position_edited.emit(position_id, {'deleted': True})
                else:
                    QMessageBox.warning(self, "Error", "Position not found")
            except Exception as e:
                self.db_session.rollback()
                QMessageBox.warning(self, "Error", f"Failed to delete position: {e}")

    # ========== View State Management (Phase 1 Custom Views) ==========

    def _on_header_context_menu(self, pos: QPoint):
        """Show context menu for column management."""
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                padding: 5px;
            }
            QMenu::item {
                color: #E0E0E0;
                padding: 8px 20px;
            }
            QMenu::item:selected {
                background-color: #0078D4;
            }
            QMenu::separator {
                height: 1px;
                background-color: #4A4A4D;
                margin: 5px 10px;
            }
        """)

        # Select columns action (opens dialog)
        select_cols_action = menu.addAction("📋 Select Columns...")
        select_cols_action.triggered.connect(self._open_column_visibility_dialog)

        menu.addSeparator()

        # Reset column order
        reset_action = menu.addAction("↺ Reset Column Order")
        reset_action.triggered.connect(self._reset_column_order)

        menu.exec(self.table_view.horizontalHeader().mapToGlobal(pos))

    def _open_column_visibility_dialog(self):
        """Open the column visibility selection dialog."""
        dialog = ColumnVisibilityDialog(
            columns=PositionTableModel.COLUMNS,
            hidden_columns=self._hidden_columns,
            parent=self
        )

        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Get new hidden columns and apply
            new_hidden = dialog.get_hidden_columns()
            self._apply_column_visibility(new_hidden)

    def _apply_column_visibility(self, hidden_columns: set):
        """Apply column visibility changes from dialog."""
        header = self.table_view.horizontalHeader()

        # Update all columns based on new hidden set
        for i in range(header.count()):
            should_hide = i in hidden_columns
            header.setSectionHidden(i, should_hide)
            if i in self._filter_buttons:
                self._filter_buttons[i].setVisible(not should_hide)

        # Update stored state
        self._hidden_columns = hidden_columns
        self._sync_filter_widths()
        self._save_view_state()

    def _toggle_column_visibility(self, logical_index: int, visible: bool):
        """Toggle visibility of a column and save state."""
        header = self.table_view.horizontalHeader()
        header.setSectionHidden(logical_index, not visible)

        # Update hidden columns set
        if visible:
            self._hidden_columns.discard(logical_index)
        else:
            self._hidden_columns.add(logical_index)

        # Update filter button visibility
        if logical_index in self._filter_buttons:
            self._filter_buttons[logical_index].setVisible(visible)

        self._sync_filter_widths()
        self._save_view_state()

    def _reset_column_order(self):
        """Reset columns to default order."""
        header = self.table_view.horizontalHeader()

        # Move each column to its default position
        for default_visual_index in range(header.count()):
            logical_index = default_visual_index  # Default: logical == visual
            current_visual_index = header.visualIndex(logical_index)
            if current_visual_index != default_visual_index:
                header.moveSection(current_visual_index, default_visual_index)

        self._column_order = list(range(len(PositionTableModel.COLUMNS)))
        self._sync_filter_widths()
        self._save_view_state()

    def _on_column_moved(self, logical_index: int, old_visual: int, new_visual: int):
        """Handle column reorder and save state."""
        # Update column order by reading current visual order
        header = self.table_view.horizontalHeader()
        self._column_order = [header.logicalIndex(visual) for visual in range(header.count())]
        self._sync_filter_widths()
        self._save_view_state()

    def _load_view_state(self):
        """Load saved views from config file and populate view combo."""
        try:
            self._loading_view = True

            if not os.path.exists(self._config_path):
                # Create default view
                self._saved_views = {"Default": self._get_current_view_state()}
                self._current_view_name = "Default"
                self._populate_view_combo()
                return

            with open(self._config_path, 'r') as f:
                config = yaml.safe_load(f) or {}

            gui_config = config.get('gui', {})
            table_views_config = gui_config.get('table_views', {})

            # Migrate from old single-view format if needed
            old_table_view = gui_config.get('table_view', {})
            if old_table_view and not table_views_config:
                # Migrate old format to new multi-view format
                self._saved_views = {
                    "Default": {
                        'hidden_columns': old_table_view.get('hidden_columns', []),
                        'column_order': old_table_view.get('column_order', list(range(len(PositionTableModel.COLUMNS)))),
                        'column_widths': {},
                        'sort_columns': []
                    }
                }
                self._current_view_name = "Default"
            else:
                # Load all saved views
                self._saved_views = {}
                for view_name, view_data in table_views_config.get('views', {}).items():
                    self._saved_views[view_name] = view_data

                # Get current view name
                self._current_view_name = table_views_config.get('current_view', 'Default')

                # Ensure at least Default view exists
                if not self._saved_views:
                    self._saved_views = {"Default": self._get_current_view_state()}
                    self._current_view_name = "Default"

            # Load the current view's state
            if self._current_view_name in self._saved_views:
                view_data = self._saved_views[self._current_view_name]
                hidden_cols = view_data.get('hidden_columns', [])
                if isinstance(hidden_cols, list):
                    self._hidden_columns = set(hidden_cols)

                col_order = view_data.get('column_order', [])
                if isinstance(col_order, list) and len(col_order) == len(PositionTableModel.COLUMNS):
                    self._column_order = col_order

                col_widths = view_data.get('column_widths', {})
                if isinstance(col_widths, dict):
                    self._column_widths = {int(k): v for k, v in col_widths.items()}

                sort_cols = view_data.get('sort_columns', [])
                if isinstance(sort_cols, list):
                    self._sort_columns = [tuple(s) for s in sort_cols]

            # Populate view combo
            self._populate_view_combo()

        except Exception as e:
            import logging
            logging.getLogger('canslim.gui').warning(f"Failed to load table view state: {e}")
        finally:
            self._loading_view = False

    def _populate_view_combo(self):
        """Populate the view combo box with saved views."""
        self.view_combo.blockSignals(True)
        self.view_combo.clear()
        for view_name in sorted(self._saved_views.keys()):
            self.view_combo.addItem(view_name)
        # Select current view
        idx = self.view_combo.findText(self._current_view_name)
        if idx >= 0:
            self.view_combo.setCurrentIndex(idx)
        self.view_combo.blockSignals(False)

    def _get_current_view_state(self) -> dict:
        """Get the current view state as a dictionary."""
        return {
            'hidden_columns': list(self._hidden_columns),
            'column_order': self._column_order.copy(),
            'column_widths': self._column_widths.copy(),
            'sort_columns': [list(s) for s in self._sort_columns]
        }

    def _save_view_state(self):
        """Save current view state to config file."""
        if self._loading_view:
            return

        try:
            # Update the current view in saved_views
            self._saved_views[self._current_view_name] = self._get_current_view_state()

            # Load existing config
            config = {}
            if os.path.exists(self._config_path):
                with open(self._config_path, 'r') as f:
                    config = yaml.safe_load(f) or {}

            # Ensure gui section exists
            if 'gui' not in config:
                config['gui'] = {}

            # Update table_views section (new multi-view format)
            config['gui']['table_views'] = {
                'current_view': self._current_view_name,
                'views': self._saved_views
            }

            # Remove old single-view format if present
            if 'table_view' in config['gui']:
                del config['gui']['table_view']

            # Save back to file
            with open(self._config_path, 'w') as f:
                yaml.dump(config, f, default_flow_style=False, sort_keys=False)

        except Exception as e:
            import logging
            logging.getLogger('canslim.gui').warning(f"Failed to save table view state: {e}")

    def _apply_view_state(self):
        """Apply loaded view state to the table."""
        self._loading_view = True
        header = self.table_view.horizontalHeader()

        # First, show all columns (reset hidden state)
        for i in range(header.count()):
            header.setSectionHidden(i, False)
            if i in self._filter_buttons:
                self._filter_buttons[i].setVisible(True)

        # Apply column order (move columns to saved positions)
        for visual_index, logical_index in enumerate(self._column_order):
            current_visual = header.visualIndex(logical_index)
            if current_visual != visual_index:
                header.moveSection(current_visual, visual_index)

        # Apply column widths
        for logical_index, width in self._column_widths.items():
            if logical_index < header.count():
                header.resizeSection(logical_index, width)

        # Apply hidden columns
        for logical_index in self._hidden_columns:
            header.setSectionHidden(logical_index, True)
            if logical_index in self._filter_buttons:
                self._filter_buttons[logical_index].setVisible(False)

        # Apply sorting
        if self._sort_columns:
            # Apply primary sort
            col_idx, ascending = self._sort_columns[0]
            order = Qt.SortOrder.AscendingOrder if ascending else Qt.SortOrder.DescendingOrder
            self.table_view.sortByColumn(col_idx, order)

        # Sync filter widths after applying state
        self._loading_view = False
        QTimer.singleShot(100, self._sync_filter_widths)

    def _on_view_changed(self, view_name: str):
        """Handle view selection change."""
        if self._loading_view or not view_name:
            return

        if view_name in self._saved_views:
            self._current_view_name = view_name
            view_data = self._saved_views[view_name]

            # Load view state
            hidden_cols = view_data.get('hidden_columns', [])
            self._hidden_columns = set(hidden_cols) if isinstance(hidden_cols, list) else set()

            col_order = view_data.get('column_order', [])
            if isinstance(col_order, list) and len(col_order) == len(PositionTableModel.COLUMNS):
                self._column_order = col_order
            else:
                self._column_order = list(range(len(PositionTableModel.COLUMNS)))

            col_widths = view_data.get('column_widths', {})
            self._column_widths = {int(k): v for k, v in col_widths.items()} if isinstance(col_widths, dict) else {}

            sort_cols = view_data.get('sort_columns', [])
            self._sort_columns = [tuple(s) for s in sort_cols] if isinstance(sort_cols, list) else []

            # Apply the view
            self._apply_view_state()

            # Save current view selection
            self._save_view_state()

    def _save_current_view(self):
        """Save changes to the current view."""
        self._save_view_state()
        QMessageBox.information(self, "View Saved", f"View '{self._current_view_name}' has been saved.")

    def _save_view_as(self):
        """Save current view with a new name."""
        name, ok = QInputDialog.getText(
            self,
            "Save View As",
            "Enter view name:",
            QLineEdit.EchoMode.Normal,
            ""
        )

        if ok and name:
            name = name.strip()
            if not name:
                QMessageBox.warning(self, "Invalid Name", "View name cannot be empty.")
                return

            # Check if name already exists
            if name in self._saved_views and name != self._current_view_name:
                reply = QMessageBox.question(
                    self,
                    "Overwrite View?",
                    f"A view named '{name}' already exists. Overwrite it?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return

            # Save as new view
            self._saved_views[name] = self._get_current_view_state()
            self._current_view_name = name
            self._populate_view_combo()
            self._save_view_state()
            QMessageBox.information(self, "View Saved", f"View '{name}' has been saved.")

    def _delete_current_view(self):
        """Delete the current view."""
        if self._current_view_name == "Default":
            QMessageBox.warning(self, "Cannot Delete", "The Default view cannot be deleted.")
            return

        reply = QMessageBox.question(
            self,
            "Delete View?",
            f"Are you sure you want to delete the view '{self._current_view_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            del self._saved_views[self._current_view_name]
            self._current_view_name = "Default"
            self._populate_view_combo()

            # Load Default view
            if "Default" in self._saved_views:
                view_data = self._saved_views["Default"]
                self._hidden_columns = set(view_data.get('hidden_columns', []))
                self._column_order = view_data.get('column_order', list(range(len(PositionTableModel.COLUMNS))))
                self._column_widths = {int(k): v for k, v in view_data.get('column_widths', {}).items()}
                self._sort_columns = [tuple(s) for s in view_data.get('sort_columns', [])]
                self._apply_view_state()

            self._save_view_state()

    def _on_column_resized(self, logical_index: int, old_size: int, new_size: int):
        """Handle column resize and save width."""
        if self._loading_view:
            return
        self._column_widths[logical_index] = new_size
        # Debounce save - don't save on every resize event
        if not hasattr(self, '_resize_timer'):
            self._resize_timer = QTimer(self)
            self._resize_timer.setSingleShot(True)
            self._resize_timer.timeout.connect(self._save_view_state)
        self._resize_timer.start(500)  # Save after 500ms of no resizing

    def _on_sort_changed(self, logical_index: int, order: Qt.SortOrder):
        """Handle sort indicator change and save sort state."""
        if self._loading_view:
            return
        ascending = order == Qt.SortOrder.AscendingOrder
        self._sort_columns = [(logical_index, ascending)]
        self._save_view_state()

    def _export_to_csv(self):
        """Export visible/filtered data to CSV file."""
        # Get file path from user
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            "",
            "CSV Files (*.csv);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith('.csv'):
            file_path += '.csv'

        try:
            header = self.table_view.horizontalHeader()

            # Get visible columns in visual order
            visible_columns = []
            for visual_idx in range(header.count()):
                logical_idx = header.logicalIndex(visual_idx)
                if not header.isSectionHidden(logical_idx):
                    visible_columns.append(logical_idx)

            # Get column names
            col_names = [PositionTableModel.COLUMNS[idx][1] for idx in visible_columns]

            # Get filtered rows from proxy model
            rows = []
            for proxy_row in range(self.proxy_model.rowCount()):
                row_data = []
                for logical_idx in visible_columns:
                    proxy_index = self.proxy_model.index(proxy_row, logical_idx)
                    display_value = self.proxy_model.data(proxy_index, Qt.ItemDataRole.DisplayRole)
                    row_data.append(display_value if display_value is not None else "")
                rows.append(row_data)

            # Write to CSV
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(col_names)
                writer.writerows(rows)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Exported {len(rows)} rows to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export data:\n{str(e)}"
            )

    def _export_to_excel(self):
        """Export visible/filtered data to Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                self,
                "Excel Export Unavailable",
                "Excel export requires the 'openpyxl' package.\n\n"
                "Install it with: pip install openpyxl"
            )
            return

        # Get file path from user
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            header = self.table_view.horizontalHeader()

            # Get visible columns in visual order
            visible_columns = []
            for visual_idx in range(header.count()):
                logical_idx = header.logicalIndex(visual_idx)
                if not header.isSectionHidden(logical_idx):
                    visible_columns.append(logical_idx)

            # Get column names
            col_names = [PositionTableModel.COLUMNS[idx][1] for idx in visible_columns]

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Positions"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

            # Write headers
            for col_idx, col_name in enumerate(col_names, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Get filtered rows from proxy model
            for proxy_row in range(self.proxy_model.rowCount()):
                for col_idx, logical_idx in enumerate(visible_columns, 1):
                    proxy_index = self.proxy_model.index(proxy_row, logical_idx)
                    display_value = self.proxy_model.data(proxy_index, Qt.ItemDataRole.DisplayRole)
                    ws.cell(row=proxy_row + 2, column=col_idx, value=display_value if display_value else "")

            # Auto-adjust column widths
            for col_idx, _ in enumerate(col_names, 1):
                max_length = max(
                    len(str(ws.cell(row=row, column=col_idx).value or ""))
                    for row in range(1, ws.max_row + 1)
                )
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Exported {self.proxy_model.rowCount()} rows to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export data:\n{str(e)}"
            )

"""
CANSLIM Monitor - Position History Dialog (Spreadsheet Comparison View)

Shows position snapshots over time in a spreadsheet format.
Each row represents the position state at a specific point in time.
Allows comparing how field values changed throughout the lifecycle.
"""

from datetime import datetime
from typing import List, Optional, Dict, Any, Set
import csv

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableView,
    QHeaderView, QPushButton, QComboBox, QFrame, QCheckBox,
    QScrollArea, QWidget, QFileDialog, QMessageBox, QMenu
)
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PyQt6.QtGui import QFont, QColor, QBrush

from canslim_monitor.data.models import TRACKED_FIELDS


# Define display names and column order for the spreadsheet
COLUMN_CONFIG = [
    # Timestamp columns always first
    ('snapshot_time', 'Snapshot Time'),
    ('change_source', 'Source'),
    # Core fields
    ('symbol', 'Symbol'),
    ('state', 'State'),
    ('portfolio', 'Portfolio'),
    ('pattern', 'Pattern'),
    # Prices
    ('pivot', 'Pivot'),
    ('stop_price', 'Stop Price'),
    ('avg_cost', 'Avg Cost'),
    # Position Management
    ('total_shares', 'Total Shares'),
    ('e1_shares', 'E1 Shares'),
    ('e1_price', 'E1 Price'),
    ('e2_shares', 'E2 Shares'),
    ('e2_price', 'E2 Price'),
    ('e3_shares', 'E3 Shares'),
    ('e3_price', 'E3 Price'),
    ('tp1_sold', 'TP1 Sold'),
    ('tp1_price', 'TP1 Price'),
    ('tp2_sold', 'TP2 Sold'),
    ('tp2_price', 'TP2 Price'),
    # Dates
    ('watch_date', 'Watch Date'),
    ('entry_date', 'Entry Date'),
    ('breakout_date', 'Breakout Date'),
    ('earnings_date', 'Earnings Date'),
    # Percentages
    ('hard_stop_pct', 'Hard Stop %'),
    ('tp1_pct', 'TP1 %'),
    ('tp2_pct', 'TP2 %'),
    # CANSLIM Ratings
    ('rs_rating', 'RS Rating'),
    ('rs_3mo', 'RS 3Mo'),
    ('rs_6mo', 'RS 6Mo'),
    ('eps_rating', 'EPS Rating'),
    ('comp_rating', 'Comp Rating'),
    ('smr_rating', 'SMR'),
    ('ad_rating', 'A/D'),
    ('industry_rank', 'Ind Rank'),
    ('fund_count', 'Fund Count'),
    ('prior_fund_count', 'Prior Fund Ct'),
    ('funds_qtr_chg', 'Funds Chg'),
    ('ud_vol_ratio', 'U/D Vol'),
    # Base characteristics
    ('base_stage', 'Base Stage'),
    ('base_depth', 'Base Depth'),
    ('base_length', 'Base Length'),
    ('prior_uptrend', 'Prior Uptrend'),
    ('breakout_vol_pct', 'BO Vol %'),
    ('breakout_price_pct', 'BO Price %'),
    # Exit info
    ('close_price', 'Close Price'),
    ('close_date', 'Close Date'),
    ('close_reason', 'Close Reason'),
    ('realized_pnl', 'Realized P&L'),
    ('realized_pnl_pct', 'Realized %'),
    # Scoring
    ('entry_grade', 'Entry Grade'),
    ('entry_score', 'Entry Score'),
    # Targets
    ('tp1_target', 'TP1 Target'),
    ('tp2_target', 'TP2 Target'),
    # State -1.5 fields
    ('original_pivot', 'Orig Pivot'),
    ('ma_test_count', 'MA Tests'),
    # Flags
    ('py1_done', 'PY1 Done'),
    ('py2_done', 'PY2 Done'),
    # Notes
    ('notes', 'Notes'),
]


class SnapshotTableModel(QAbstractTableModel):
    """
    Table model for displaying position snapshots over time.
    Each row is a snapshot of the position at a specific timestamp.
    """

    def __init__(self, snapshots: List[Dict[str, Any]], parent=None):
        super().__init__(parent)
        self._snapshots = snapshots
        self._columns = COLUMN_CONFIG
        self._hidden_columns: Set[int] = set()

    def rowCount(self, parent=QModelIndex()):
        return len(self._snapshots)

    def columnCount(self, parent=QModelIndex()):
        return len(self._columns)

    def data(self, index: QModelIndex, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None

        row = index.row()
        col = index.column()

        if row >= len(self._snapshots) or col >= len(self._columns):
            return None

        snapshot = self._snapshots[row]
        field_key = self._columns[col][0]
        value = snapshot.get(field_key)

        if role == Qt.ItemDataRole.DisplayRole:
            return self._format_value(value, field_key)

        elif role == Qt.ItemDataRole.ForegroundRole:
            # Highlight changed values (not the first row/current)
            if row == 0:
                return QBrush(QColor("#4FC3F7"))  # Current row in light blue
            # Check if this field changed from prior row
            if row < len(self._snapshots) - 1:
                next_snapshot = self._snapshots[row + 1]
                next_value = next_snapshot.get(field_key)
                if self._values_different(value, next_value):
                    return QBrush(QColor("#FFD54F"))  # Changed values in yellow
            return QBrush(QColor("#CCCCCC"))

        elif role == Qt.ItemDataRole.BackgroundRole:
            # Alternating row colors
            if row % 2 == 0:
                return QBrush(QColor("#252526"))
            else:
                return QBrush(QColor("#2D2D30"))

        elif role == Qt.ItemDataRole.TextAlignmentRole:
            # Right-align numeric columns
            if field_key in {'pivot', 'stop_price', 'avg_cost', 'e1_price', 'e2_price',
                           'e3_price', 'tp1_price', 'tp2_price', 'close_price',
                           'realized_pnl', 'tp1_target', 'tp2_target', 'original_pivot'}:
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            if field_key in {'total_shares', 'e1_shares', 'e2_shares', 'e3_shares',
                           'tp1_sold', 'tp2_sold', 'rs_rating', 'rs_3mo', 'rs_6mo',
                           'eps_rating', 'comp_rating', 'industry_rank', 'fund_count',
                           'prior_fund_count', 'funds_qtr_chg', 'base_length', 'entry_score', 'ma_test_count'}:
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            return Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter

        return None

    def headerData(self, section: int, orientation: Qt.Orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                if section < len(self._columns):
                    return self._columns[section][1]
            else:
                if section == 0:
                    return "Current"
                return f"T-{section}"
        return None

    def _format_value(self, value: Any, field_key: str) -> str:
        """Format a value for display based on field type."""
        if value is None:
            return "—"

        # Price fields
        price_fields = {
            'pivot', 'stop_price', 'avg_cost', 'e1_price', 'e2_price', 'e3_price',
            'tp1_price', 'tp2_price', 'close_price', 'tp1_target', 'tp2_target',
            'original_pivot', 'realized_pnl'
        }

        # Percentage fields
        pct_fields = {
            'hard_stop_pct', 'tp1_pct', 'tp2_pct', 'base_depth',
            'breakout_vol_pct', 'breakout_price_pct', 'prior_uptrend',
            'realized_pnl_pct', 'ud_vol_ratio'
        }

        # Integer fields
        int_fields = {
            'e1_shares', 'e2_shares', 'e3_shares', 'total_shares',
            'tp1_sold', 'tp2_sold', 'rs_rating', 'rs_3mo', 'rs_6mo',
            'eps_rating', 'comp_rating', 'industry_rank', 'fund_count',
            'funds_qtr_chg', 'base_length', 'entry_score', 'ma_test_count'
        }

        # State formatting
        if field_key == 'state':
            return self._format_state(value)

        # Source formatting
        if field_key == 'change_source':
            return self._format_source(value)

        # Date formatting
        if field_key in {'snapshot_time', 'watch_date', 'entry_date', 'breakout_date',
                        'earnings_date', 'close_date', 'e1_date', 'e2_date', 'e3_date',
                        'tp1_date', 'tp2_date'}:
            return self._format_date(value)

        # Boolean fields
        if field_key in {'py1_done', 'py2_done'}:
            if isinstance(value, bool):
                return "Yes" if value else "No"
            if isinstance(value, str):
                return "Yes" if value.lower() in ('true', 'yes', '1') else "No"
            return str(value)

        try:
            if field_key in price_fields:
                num = float(value)
                return f"${num:,.2f}"
            elif field_key in pct_fields:
                num = float(value)
                return f"{num:.1f}%"
            elif field_key in int_fields:
                return str(int(float(value)))
            else:
                return str(value)
        except (ValueError, TypeError):
            return str(value)

    def _format_state(self, value: Any) -> str:
        """Format state value with friendly name."""
        try:
            state = float(value)
            state_names = {
                -2: "Stopped",
                -1.5: "Exit Watch",
                -1: "Closed",
                0: "Watching",
                1: "Entry 1",
                2: "Entry 2",
                3: "Entry 3",
                4: "TP1",
                5: "TP2",
            }
            return state_names.get(state, str(state))
        except (ValueError, TypeError):
            return str(value)

    def _format_source(self, value: str) -> str:
        """Format change source for display."""
        if not value:
            return "—"
        source_display = {
            'manual_edit': 'Manual',
            'state_transition': 'State',
            'system_calc': 'System',
            'price_update': 'Price',
            'current': 'Current',
        }
        return source_display.get(value, value)

    def _format_date(self, value: Any) -> str:
        """Format date/datetime for display."""
        if not value:
            return "—"

        if isinstance(value, str):
            try:
                dt = datetime.fromisoformat(value)
                return dt.strftime("%Y-%m-%d %H:%M")
            except ValueError:
                return value

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M")

        return str(value)

    def _values_different(self, val1: Any, val2: Any) -> bool:
        """Check if two values are different for highlighting."""
        if val1 is None and val2 is None:
            return False
        if val1 is None or val2 is None:
            return True
        try:
            if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
                return abs(float(val1) - float(val2)) > 0.0001
        except (ValueError, TypeError):
            pass
        return str(val1) != str(val2)

    def get_column_key(self, col: int) -> str:
        """Get the field key for a column index."""
        if col < len(self._columns):
            return self._columns[col][0]
        return ""


class ColumnVisibilityDialog(QDialog):
    """Dialog for selecting which columns to show/hide."""

    def __init__(self, columns: list, hidden_columns: set, parent=None):
        super().__init__(parent)
        self.columns = columns
        self.hidden_columns = set(hidden_columns)
        self._checkboxes: Dict[int, QCheckBox] = {}

        self.setWindowTitle("Select Columns")
        self.setMinimumWidth(300)
        self.setMaximumHeight(600)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        self.setStyleSheet("""
            QDialog {
                background-color: #2D2D30;
            }
            QLabel {
                color: #E0E0E0;
            }
            QCheckBox {
                color: #E0E0E0;
                padding: 4px 8px;
                font-size: 12px;
            }
            QCheckBox:hover {
                background-color: #3C3C3F;
                border-radius: 3px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
            QCheckBox::indicator:checked {
                background-color: #0078D4;
                border: 1px solid #0078D4;
                border-radius: 3px;
            }
            QCheckBox::indicator:unchecked {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                border-radius: 3px;
            }
            QPushButton {
                background-color: #3C3C3F;
                border: 1px solid #4A4A4D;
                border-radius: 4px;
                padding: 6px 12px;
                color: #E0E0E0;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #4A4A4D;
                border-color: #0078D4;
            }
            QScrollArea {
                border: 1px solid #4A4A4D;
                border-radius: 4px;
                background-color: #2D2D30;
            }
        """)

        # Title
        title = QLabel("Select Columns to Display")
        title.setStyleSheet("font-weight: bold; font-size: 13px; padding: 2px;")
        layout.addWidget(title)

        # Select All / Select None buttons
        btn_layout = QHBoxLayout()
        select_all_btn = QPushButton("Select All")
        select_all_btn.clicked.connect(self._select_all)
        btn_layout.addWidget(select_all_btn)

        select_none_btn = QPushButton("Select None")
        select_none_btn.clicked.connect(self._select_none)
        btn_layout.addWidget(select_none_btn)
        layout.addLayout(btn_layout)

        # Scrollable checkbox list
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(400)

        list_widget = QWidget()
        list_widget.setStyleSheet("background-color: #2D2D30;")
        list_layout = QVBoxLayout(list_widget)
        list_layout.setContentsMargins(4, 4, 4, 4)
        list_layout.setSpacing(2)

        for idx, (key, name) in enumerate(self.columns):
            cb = QCheckBox(name)
            cb.setChecked(idx not in self.hidden_columns)
            self._checkboxes[idx] = cb
            list_layout.addWidget(cb)

        list_layout.addStretch()
        scroll.setWidget(list_widget)
        layout.addWidget(scroll)

        # Apply / Cancel buttons
        action_layout = QHBoxLayout()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        action_layout.addWidget(cancel_btn)

        action_layout.addStretch()

        apply_btn = QPushButton("Apply")
        apply_btn.clicked.connect(self.accept)
        apply_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        action_layout.addWidget(apply_btn)
        layout.addLayout(action_layout)

    def _select_all(self):
        for cb in self._checkboxes.values():
            cb.setChecked(True)

    def _select_none(self):
        for idx, cb in self._checkboxes.items():
            # Keep first two columns (Timestamp, Source) always visible
            if idx < 2:
                cb.setChecked(True)
            else:
                cb.setChecked(False)

    def get_hidden_columns(self) -> set:
        hidden = set()
        for idx, cb in self._checkboxes.items():
            if not cb.isChecked():
                hidden.add(idx)
        return hidden


class PositionHistoryDialog(QDialog):
    """
    Dialog showing position history as a spreadsheet comparison.

    Each row represents a snapshot of the position at a point in time,
    allowing easy comparison of how fields changed over the lifecycle.
    """

    def __init__(
        self,
        symbol: str,
        position_id: int,
        current_position: Dict[str, Any],
        history: List[Dict[str, Any]],
        parent=None
    ):
        super().__init__(parent)
        self.symbol = symbol
        self.position_id = position_id
        self.current_position = current_position
        self.raw_history = history

        self._hidden_columns: Set[int] = set()
        self._snapshots: List[Dict[str, Any]] = []

        self._build_snapshots()
        self._setup_ui()

    def _build_snapshots(self):
        """
        Build snapshots by reconstructing position state at each change time.

        Starting from current values, we go backwards through history and
        reconstruct what the position looked like at each change point.
        """
        # Start with current position as first snapshot
        current_snapshot = {
            'snapshot_time': datetime.now(),
            'change_source': 'current',
        }

        # Copy all tracked fields from current position
        for key, _ in COLUMN_CONFIG:
            if key not in ('snapshot_time', 'change_source'):
                current_snapshot[key] = self.current_position.get(key)

        self._snapshots = [current_snapshot]

        # Group history by timestamp (changes made at same time)
        if not self.raw_history:
            return

        # Sort history by changed_at descending (most recent first)
        sorted_history = sorted(
            self.raw_history,
            key=lambda x: x.get('changed_at', datetime.min),
            reverse=True
        )

        # Build snapshot for each change batch
        current_values = dict(current_snapshot)
        last_timestamp = None

        for entry in sorted_history:
            changed_at = entry.get('changed_at')
            field_name = entry.get('field_name')
            old_value = entry.get('old_value')
            new_value = entry.get('new_value')
            change_source = entry.get('change_source')

            # If this is a new timestamp, save the previous snapshot
            if last_timestamp is None or changed_at != last_timestamp:
                if last_timestamp is not None:
                    # Create snapshot with the state BEFORE this change
                    snapshot = dict(current_values)
                    snapshot['snapshot_time'] = last_timestamp
                    snapshot['change_source'] = last_source
                    self._snapshots.append(snapshot)

                last_timestamp = changed_at
                last_source = change_source

            # Revert this field to its old value
            if field_name in current_values:
                current_values[field_name] = old_value

        # Add final snapshot (oldest state)
        if last_timestamp:
            snapshot = dict(current_values)
            snapshot['snapshot_time'] = last_timestamp
            snapshot['change_source'] = last_source
            self._snapshots.append(snapshot)

    def _setup_ui(self):
        """Set up the dialog UI."""
        self.setWindowTitle(f"Position History: {self.symbol}")
        self.setMinimumSize(1000, 600)

        # Apply dark theme
        self.setStyleSheet("""
            QDialog {
                background-color: #1E1E1E;
            }
            QLabel {
                color: #CCCCCC;
            }
            QTableView {
                background-color: #252526;
                alternate-background-color: #2D2D30;
                color: #CCCCCC;
                gridline-color: #3F3F46;
                border: 1px solid #3F3F46;
                border-radius: 4px;
                selection-background-color: #0078D4;
            }
            QHeaderView::section {
                background-color: #2D2D30;
                color: #CCCCCC;
                padding: 6px;
                border: none;
                border-bottom: 1px solid #3F3F46;
                border-right: 1px solid #3F3F46;
                font-weight: bold;
            }
            QPushButton {
                background-color: #3C3C3C;
                color: #CCCCCC;
                border: 1px solid #3F3F46;
                border-radius: 4px;
                padding: 6px 16px;
            }
            QPushButton:hover {
                background-color: #4D4D4D;
                border-color: #0078D4;
            }
            QMenu {
                background-color: #2D2D30;
                color: #CCCCCC;
                border: 1px solid #3F3F46;
            }
            QMenu::item:selected {
                background-color: #0078D4;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Header
        header_layout = QHBoxLayout()

        title = QLabel(f"Position History: {self.symbol}")
        title_font = QFont()
        title_font.setBold(True)
        title_font.setPointSize(14)
        title.setFont(title_font)
        header_layout.addWidget(title)

        header_layout.addStretch()

        # Column visibility button
        columns_btn = QPushButton("Columns...")
        columns_btn.clicked.connect(self._show_column_dialog)
        header_layout.addWidget(columns_btn)

        # Export button with menu
        export_btn = QPushButton("Export")
        export_menu = QMenu(self)
        export_csv_action = export_menu.addAction("Export to CSV")
        export_csv_action.triggered.connect(self._export_to_csv)
        export_excel_action = export_menu.addAction("Export to Excel")
        export_excel_action.triggered.connect(self._export_to_excel)
        export_btn.setMenu(export_menu)
        header_layout.addWidget(export_btn)

        layout.addLayout(header_layout)

        # Summary
        summary_text = f"Showing {len(self._snapshots)} snapshots"
        if len(self.raw_history) > 0:
            summary_text += f" ({len(self.raw_history)} field changes recorded)"
        summary_label = QLabel(summary_text)
        summary_label.setStyleSheet("color: #888888;")
        layout.addWidget(summary_label)

        # Legend
        legend_layout = QHBoxLayout()
        legend_layout.addWidget(QLabel("Legend:"))

        current_label = QLabel("Current")
        current_label.setStyleSheet("color: #4FC3F7; font-weight: bold; padding: 0 8px;")
        legend_layout.addWidget(current_label)

        changed_label = QLabel("Changed value")
        changed_label.setStyleSheet("color: #FFD54F; font-weight: bold; padding: 0 8px;")
        legend_layout.addWidget(changed_label)

        legend_layout.addStretch()
        layout.addLayout(legend_layout)

        # Table
        self.model = SnapshotTableModel(self._snapshots)
        self.table_view = QTableView()
        self.table_view.setModel(self.model)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table_view.setSortingEnabled(False)
        self.table_view.verticalHeader().setVisible(True)

        # Configure header
        header = self.table_view.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setDefaultSectionSize(100)

        # Set initial column widths
        for i in range(self.model.columnCount()):
            key = self.model.get_column_key(i)
            if key == 'snapshot_time':
                header.resizeSection(i, 140)
            elif key == 'notes':
                header.resizeSection(i, 200)
            elif key in {'symbol', 'portfolio', 'pattern', 'close_reason'}:
                header.resizeSection(i, 100)
            elif key in {'state', 'change_source'}:
                header.resizeSection(i, 80)

        layout.addWidget(self.table_view)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1084D8;
            }
        """)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)

        # Apply initial column visibility
        self._apply_column_visibility()

    def _show_column_dialog(self):
        """Show column visibility dialog."""
        dialog = ColumnVisibilityDialog(
            COLUMN_CONFIG,
            self._hidden_columns,
            self
        )
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._hidden_columns = dialog.get_hidden_columns()
            self._apply_column_visibility()

    def _apply_column_visibility(self):
        """Apply column visibility settings to the table."""
        header = self.table_view.horizontalHeader()
        for col in range(self.model.columnCount()):
            header.setSectionHidden(col, col in self._hidden_columns)

    def _export_to_csv(self):
        """Export visible data to CSV file."""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            f"{self.symbol}_history.csv",
            "CSV Files (*.csv);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith('.csv'):
            file_path += '.csv'

        try:
            header = self.table_view.horizontalHeader()

            # Get visible columns
            visible_columns = []
            for col in range(self.model.columnCount()):
                if not header.isSectionHidden(col):
                    visible_columns.append(col)

            # Get column names
            col_names = [COLUMN_CONFIG[col][1] for col in visible_columns]

            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(col_names)

                for row in range(self.model.rowCount()):
                    row_data = []
                    for col in visible_columns:
                        index = self.model.index(row, col)
                        value = self.model.data(index, Qt.ItemDataRole.DisplayRole)
                        row_data.append(value if value else "")
                    writer.writerow(row_data)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Exported {self.model.rowCount()} rows to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export data:\n{str(e)}"
            )

    def _export_to_excel(self):
        """Export visible data to Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                self,
                "Excel Export Unavailable",
                "Excel export requires the 'openpyxl' package.\n\n"
                "Install it with: pip install openpyxl"
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            f"{self.symbol}_history.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            header = self.table_view.horizontalHeader()

            # Get visible columns
            visible_columns = []
            for col in range(self.model.columnCount()):
                if not header.isSectionHidden(col):
                    visible_columns.append(col)

            # Get column names
            col_names = [COLUMN_CONFIG[col][1] for col in visible_columns]

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.symbol} History"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

            # Write headers
            for col_idx, col_name in enumerate(col_names, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row in range(self.model.rowCount()):
                for col_idx, logical_col in enumerate(visible_columns, 1):
                    index = self.model.index(row, logical_col)
                    value = self.model.data(index, Qt.ItemDataRole.DisplayRole)
                    ws.cell(row=row + 2, column=col_idx, value=value if value else "")

            # Auto-adjust column widths
            for col_idx, _ in enumerate(col_names, 1):
                max_length = max(
                    len(str(ws.cell(row=row, column=col_idx).value or ""))
                    for row in range(1, ws.max_row + 1)
                )
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Exported {self.model.rowCount()} rows to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export data:\n{str(e)}"
            )

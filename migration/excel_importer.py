"""
CANSLIM Monitor - Excel Position Manager Importer
Migrates data from Excel Position Manager to SQLite database.
"""

import logging
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Dict, List, Any

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from canslim_monitor.data.database import DatabaseManager, init_database
from canslim_monitor.data.models import Position, MarketRegime
from canslim_monitor.data.repositories import RepositoryManager


# Column mapping: Excel column index (1-based) to database field
COLUMN_MAP = {
    1: 'portfolio',
    2: 'symbol',
    3: 'state',
    4: 'pattern',
    5: 'pivot',
    6: 'e1_shares',
    7: 'e1_price',
    8: 'e2_shares',
    9: 'e2_price',
    10: 'e3_shares',
    11: 'e3_price',
    12: 'tp1_sold',
    13: 'tp1_price',
    14: 'tp2_sold',
    15: 'tp2_price',
    16: 'hard_stop_pct',
    17: 'tp1_pct',
    18: 'tp2_pct',
    19: 'watch_date',
    20: 'entry_date',
    21: 'breakout_date',
    22: 'base_stage',
    23: 'base_depth',
    24: 'base_length',
    25: 'rs_rating',
    26: 'comp_rating',
    27: 'eps_rating',
    28: 'ad_rating',
    29: 'ud_vol_ratio',
    30: 'group_rank',
    31: 'fund_count',
    32: 'prior_uptrend',
    33: 'py1_done',
    34: 'py2_done',
    35: 'earnings_date',
    36: 'notes',
    37: 'total_shares',
    38: 'avg_cost',
    39: 'stop_price',
    40: 'tp1_target',
    41: 'tp2_target',
}

# Valid patterns from Patterns sheet
VALID_PATTERNS = [
    'Cup & Handle',
    'Cup w/Handle', 
    'Cup',
    'Flat Base',
    'Double Bottom',
    'Double Botton',  # Common typo
    'High Tight Flag',
    'Ascending Base',
    'Saucer Base',
    'IPO Base',
    'Base on Base',
    'Consolidation',
]

# State definitions
STATE_NAMES = {
    -2: 'STOPPED_OUT',
    -1: 'CLOSED',
    0: 'WATCHING',
    1: 'ENTRY_1',
    2: 'ENTRY_2', 
    3: 'FULL_POSITION',
    4: 'TP1_HIT',
    5: 'TP2_HIT',
    6: 'TRAILING',
}


class ExcelImporter:
    """Import positions from Excel Position Manager spreadsheet."""
    
    def __init__(
        self,
        excel_path: str,
        db_path: str,
        logger: Optional[logging.Logger] = None
    ):
        """
        Initialize importer.
        
        Args:
            excel_path: Path to Excel file
            db_path: Path to SQLite database
            logger: Logger instance
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        self.excel_path = Path(excel_path)
        self.db_path = db_path
        self.logger = logger or logging.getLogger('migration')
        
        self.workbook = None
        self.db: Optional[DatabaseManager] = None
        
        # Import statistics
        self.stats = {
            'total_rows': 0,
            'imported': 0,
            'skipped': 0,
            'errors': 0,
            'by_state': {},
        }
    
    def run(self, clear_existing: bool = False) -> Dict[str, Any]:
        """
        Run the import process.
        
        Args:
            clear_existing: If True, clear existing positions before import
            
        Returns:
            Import statistics
        """
        self.logger.info(f"Starting import from {self.excel_path}")
        
        # Load Excel
        self._load_excel()
        
        # Initialize database
        self._init_database(clear_existing)
        
        # Import positions
        self._import_positions()
        
        # Import reference data
        self._import_patterns()
        
        self.logger.info(f"Import complete: {self.stats['imported']} positions imported")
        return self.stats
    
    def _load_excel(self):
        """Load Excel workbook."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        self.workbook = openpyxl.load_workbook(self.excel_path)
        self.logger.info(f"Loaded workbook with sheets: {self.workbook.sheetnames}")
    
    def _init_database(self, clear_existing: bool):
        """Initialize database."""
        self.db = init_database(self.db_path)
        self.logger.info(f"Database initialized: {self.db_path}")
        
        if clear_existing:
            session = self.db.get_new_session()
            try:
                # Must delete in correct order due to foreign key constraints
                # Delete child tables first, then positions
                from canslim_monitor.data.models import Alert, DailySnapshot, Outcome
                
                alerts_deleted = session.query(Alert).delete()
                snapshots_deleted = session.query(DailySnapshot).delete()
                outcomes_deleted = session.query(Outcome).delete()
                positions_deleted = session.query(Position).delete()
                
                session.commit()
                self.logger.info(
                    f"Cleared existing data: {positions_deleted} positions, "
                    f"{alerts_deleted} alerts, {snapshots_deleted} snapshots, "
                    f"{outcomes_deleted} outcomes"
                )
            finally:
                session.close()
    
    def _import_positions(self):
        """Import positions from Positions sheet."""
        ws = self.workbook['Positions']
        session = self.db.get_new_session()
        
        # Track symbols we've seen (keep most recent/active state)
        seen_symbols = {}  # symbol -> (row_num, state)
        
        try:
            repos = RepositoryManager(session)
            
            # First pass: identify duplicates, prefer active positions
            for row_num in range(2, ws.max_row + 1):
                position_data = self._extract_row(ws, row_num)
                if not position_data:
                    continue
                
                position_data = self._clean_position_data(position_data)
                symbol = position_data.get('symbol')
                if not symbol:
                    continue
                
                state = position_data.get('state', 0)
                portfolio = position_data.get('portfolio', 'CWB')
                key = f"{symbol}:{portfolio}"
                
                if key in seen_symbols:
                    existing_row, existing_state = seen_symbols[key]
                    # Prefer active positions (state >= 0) over closed (state < 0)
                    # Among active, prefer higher state (more progressed)
                    if state >= 0 and (existing_state < 0 or state > existing_state):
                        seen_symbols[key] = (row_num, state)
                        self.logger.debug(f"Replacing {symbol} row {existing_row} (state {existing_state}) with row {row_num} (state {state})")
                else:
                    seen_symbols[key] = (row_num, state)
            
            # Get rows to import
            rows_to_import = {row_num for row_num, _ in seen_symbols.values()}
            self.logger.info(f"Found {len(rows_to_import)} unique positions to import")
            
            # Second pass: import selected rows
            for row_num in range(2, ws.max_row + 1):
                if row_num not in rows_to_import:
                    continue
                    
                try:
                    position_data = self._extract_row(ws, row_num)
                    
                    if not position_data:
                        continue
                    
                    self.stats['total_rows'] += 1
                    
                    # Validate and clean data
                    position_data = self._clean_position_data(position_data)
                    
                    if not position_data.get('symbol'):
                        self.stats['skipped'] += 1
                        continue
                    
                    # Calculate derived fields (avg_cost, total_shares)
                    position_data = self._calculate_derived_fields(position_data)
                    
                    # Store row number for sync tracking
                    position_data['sheet_row_id'] = str(row_num)
                    
                    # Create position
                    position = repos.positions.create(**position_data)
                    
                    self.stats['imported'] += 1
                    state = position_data.get('state', 0)
                    self.stats['by_state'][state] = self.stats['by_state'].get(state, 0) + 1
                    
                    if self.stats['imported'] % 20 == 0:
                        self.logger.info(f"Imported {self.stats['imported']} positions...")
                    
                except Exception as e:
                    self.logger.error(f"Error importing row {row_num}: {e}")
                    self.stats['errors'] += 1
                    session.rollback()  # Rollback to continue with next row
            
            session.commit()
            
        finally:
            session.close()
    
    def _extract_row(self, ws, row_num: int) -> Optional[Dict]:
        """Extract data from a single row."""
        data = {}
        
        for col_num, field_name in COLUMN_MAP.items():
            cell = ws.cell(row=row_num, column=col_num)
            value = cell.value
            
            # Skip empty cells
            if value is None or (isinstance(value, str) and value.strip() == ''):
                continue
            
            data[field_name] = value
        
        return data if data else None
    
    def _clean_position_data(self, data: Dict) -> Dict:
        """Clean and validate position data."""
        cleaned = {}
        
        for field, value in data.items():
            # Handle symbol
            if field == 'symbol':
                if isinstance(value, str):
                    cleaned[field] = value.strip().upper()
                continue
            
            # Handle state
            if field == 'state':
                try:
                    cleaned[field] = int(float(value))
                except:
                    cleaned[field] = 0
                continue
            
            # Handle dates
            if field in ('watch_date', 'entry_date', 'breakout_date', 'earnings_date'):
                if isinstance(value, datetime):
                    cleaned[field] = value.date()
                elif isinstance(value, date):
                    cleaned[field] = value
                elif isinstance(value, str):
                    try:
                        cleaned[field] = datetime.strptime(value, '%Y-%m-%d').date()
                    except:
                        pass
                continue
            
            # Handle numeric fields
            if field in ('pivot', 'stop_price', 'avg_cost', 'tp1_target', 'tp2_target',
                        'e1_price', 'e2_price', 'e3_price', 'tp1_price', 'tp2_price',
                        'hard_stop_pct', 'tp1_pct', 'tp2_pct', 'base_depth', 
                        'ud_vol_ratio', 'prior_uptrend'):
                try:
                    cleaned[field] = float(value)
                except:
                    pass
                continue
            
            # Handle integer fields
            if field in ('e1_shares', 'e2_shares', 'e3_shares', 'tp1_sold', 'tp2_sold',
                        'total_shares', 'rs_rating', 'comp_rating', 'eps_rating',
                        'group_rank', 'fund_count', 'base_length'):
                try:
                    cleaned[field] = int(float(value))
                except:
                    pass
                continue
            
            # Handle boolean fields
            if field in ('py1_done', 'py2_done'):
                if isinstance(value, bool):
                    cleaned[field] = value
                elif isinstance(value, (int, float)):
                    cleaned[field] = bool(value)
                elif isinstance(value, str):
                    cleaned[field] = value.lower() in ('true', 'yes', '1', 'y')
                continue
            
            # Handle pattern validation
            if field == 'pattern':
                if isinstance(value, str):
                    pattern = value.strip()
                    # Normalize common typos
                    if pattern == 'Double Botton':
                        pattern = 'Double Bottom'
                    cleaned[field] = pattern
                continue
            
            # Handle string fields
            if field in ('portfolio', 'notes', 'base_stage', 'ad_rating'):
                if isinstance(value, str):
                    cleaned[field] = value.strip()
                else:
                    cleaned[field] = str(value)
                continue
            
            # Pass through anything else
            cleaned[field] = value
        
        return cleaned
    
    def _calculate_derived_fields(self, data: Dict) -> Dict:
        """Calculate avg_cost and total_shares from entry data."""
        e1_shares = data.get('e1_shares') or 0
        e2_shares = data.get('e2_shares') or 0
        e3_shares = data.get('e3_shares') or 0
        
        e1_price = data.get('e1_price') or 0
        e2_price = data.get('e2_price') or 0
        e3_price = data.get('e3_price') or 0
        
        # Convert to proper types
        try:
            e1_shares = int(float(e1_shares)) if e1_shares else 0
            e2_shares = int(float(e2_shares)) if e2_shares else 0
            e3_shares = int(float(e3_shares)) if e3_shares else 0
            e1_price = float(e1_price) if e1_price else 0
            e2_price = float(e2_price) if e2_price else 0
            e3_price = float(e3_price) if e3_price else 0
        except:
            return data
        
        # Calculate total shares
        total_shares = e1_shares + e2_shares + e3_shares
        if total_shares > 0:
            data['total_shares'] = total_shares
        
        # Calculate weighted average cost
        total_cost = (e1_shares * e1_price) + (e2_shares * e2_price) + (e3_shares * e3_price)
        if total_shares > 0 and total_cost > 0:
            data['avg_cost'] = total_cost / total_shares
        
        return data
    
    def _import_patterns(self):
        """Import valid patterns from Patterns sheet."""
        if 'Patterns' not in self.workbook.sheetnames:
            return
        
        ws = self.workbook['Patterns']
        patterns = []
        
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if value and isinstance(value, str) and value.strip():
                patterns.append(value.strip())
        
        self.logger.info(f"Found {len(patterns)} valid patterns: {patterns}")
    
    def get_summary(self) -> str:
        """Get import summary as formatted string."""
        lines = [
            "=" * 50,
            "IMPORT SUMMARY",
            "=" * 50,
            f"Total rows processed: {self.stats['total_rows']}",
            f"Successfully imported: {self.stats['imported']}",
            f"Skipped (empty/invalid): {self.stats['skipped']}",
            f"Errors: {self.stats['errors']}",
            "",
            "By State:",
        ]
        
        for state, count in sorted(self.stats['by_state'].items()):
            name = STATE_NAMES.get(state, f'STATE_{state}')
            lines.append(f"  {state:2} ({name:15}): {count}")
        
        lines.append("=" * 50)
        return "\n".join(lines)


def import_from_excel(
    excel_path: str,
    db_path: str,
    clear_existing: bool = False
) -> Dict[str, Any]:
    """
    Convenience function to import from Excel.
    
    Args:
        excel_path: Path to Excel file
        db_path: Path to database
        clear_existing: Clear existing data before import
        
    Returns:
        Import statistics
    """
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s'
    )
    
    importer = ExcelImporter(excel_path, db_path)
    stats = importer.run(clear_existing=clear_existing)
    print(importer.get_summary())
    return stats


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python excel_importer.py <excel_file> <database_file> [--clear]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    db_file = sys.argv[2]
    clear = '--clear' in sys.argv
    
    import_from_excel(excel_file, db_file, clear_existing=clear)
